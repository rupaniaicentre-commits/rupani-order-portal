#!/usr/bin/env python3
"""
Comprehensive audit and rebuild of Rupani Automobiles product catalog.
Tasks 1-6: Extract SAI, Extract Aerostar, Match, Re-extract images, Rebuild XLSX, Update JSON.
"""

import pdfplumber
import fitz  # PyMuPDF
import openpyxl
import json
import re
import os
import io
from pathlib import Path
from difflib import SequenceMatcher

# ─── Paths ────────────────────────────────────────────────────────────────────
SAI_PDF     = "/Users/rupaniai/Downloads/AEROSTAR/SAI ABS 2026 PRODUCT CATALOGUE.pdf"
AS_PDF      = "/Users/rupaniai/Downloads/AEROSTAR/Aerostar-ABS Price List 12th May.pdf"
CATALOG_XL  = "/Users/rupaniai/my-new-personal-project/rupani-order-portal/product_catalog.xlsx"
IMAGES_DIR  = Path("/Users/rupaniai/my-new-personal-project/rupani-order-portal/static/images")
JSON_OUT    = "/Users/rupaniai/my-new-personal-project/rupani-order-portal/aerostar_products.json"
IMAGES_DIR.mkdir(parents=True, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 1 – Extract SAI products
# ═══════════════════════════════════════════════════════════════════════════════
def extract_sai_products():
    """Extract all SAI products from the 286-page catalogue PDF."""
    print("\n═══ TASK 1: Extracting SAI products ═══")

    # Pattern: SAI-XXX (with optional letters/suffix)
    SAI_PATTERN = re.compile(r'\bSAI-[\w\d]+(?:-BP)?\b', re.IGNORECASE)

    # Known category keywords that appear as section headers
    CATEGORY_KEYWORDS = [
        "FRONT FENDER", "HEAD LIGHT VISOR", "HEADLIGHT VISOR",
        "SIDE COWL", "REAR COWL CENTER PLATE", "REAR COWL BACK PLATE",
        "REAR COWL", "T.P.F.C", "TPFC", "ENGINE GUARD", "FUEL TANK COVER",
        "NOSE", "LOWER BODY COVER", "FLOORING UNDER COVER", "FRONT BODY COVER",
        "FOOT TRIM", "FOOT REST", "MUDGUARD", "TAIL PIECE", "SAREE GUARD",
        "BATTERY COVER", "CHAIN COVER", "SILENCER COVER", "SWING ARM COVER",
        "GRAB RAIL", "LEG GUARD", "DASH BOARD", "DASHBOARD", "METER VISOR",
        "HANDLE COVER", "INNER PANEL", "OUTER PANEL", "STEP BOARD",
        "CARRIER PLATE", "REAR CARRIER", "BACK PANEL", "AIR FILTER",
        "TOOL BOX", "BODY PANEL", "WIND SHIELD", "WIND SCREEN",
        "FAIRING", "COWL", "FENDER",
    ]

    products = []
    current_category = "UNKNOWN"
    current_brand = "SAI"

    with pdfplumber.open(SAI_PDF) as pdf:
        total_pages = len(pdf.pages)
        print(f"  SAI PDF has {total_pages} pages")

        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if not text:
                continue

            lines = [l.strip() for l in text.split('\n') if l.strip()]

            # Detect category from this page's text
            for kw in CATEGORY_KEYWORDS:
                if kw in text.upper():
                    current_category = kw
                    break

            # Now extract part numbers + descriptions
            # Strategy: find lines containing SAI-XXX patterns.
            # The description is usually on the same line or the next line.
            i = 0
            while i < len(lines):
                line = lines[i]
                matches = SAI_PATTERN.findall(line)

                if matches:
                    # Split the line by SAI part numbers to extract descriptions
                    # Typically the line looks like:
                    # "SAI-65D-BP SAI-65E"  (part numbers only)
                    # or "SAI-65D Front Fender Fit For..."

                    # Check if ONLY part numbers on this line
                    stripped = SAI_PATTERN.sub('', line).strip()
                    if not stripped:
                        # Next line(s) should have descriptions
                        # Get descriptions from next lines
                        desc_lines = []
                        j = i + 1
                        while j < len(lines) and j < i + len(matches) + 2:
                            next_line = lines[j]
                            # If next line starts with SAI-, stop
                            if SAI_PATTERN.match(next_line):
                                break
                            # If it looks like a description
                            if not next_line.startswith('www.') and not next_line.isdigit():
                                desc_lines.append(next_line)
                            j += 1

                        # Pair each part number with a description chunk
                        # The descriptions are typically interleaved:
                        # desc for part1, desc for part2, etc.
                        # Split desc_lines into groups for each match
                        descs = []
                        if desc_lines:
                            # Try to split desc_lines equally among matches
                            per_match = max(1, len(desc_lines) // len(matches))
                            for idx in range(len(matches)):
                                start = idx * per_match
                                end = start + per_match
                                descs.append(' '.join(desc_lines[start:end]).strip())

                        for idx, pn in enumerate(matches):
                            desc = descs[idx] if idx < len(descs) else ''
                            products.append({
                                'sai_part_number': pn.upper(),
                                'description': desc,
                                'category': current_category,
                                'brand': 'SAI',
                                'source_page': page_num + 1,
                            })
                    else:
                        # Part numbers mixed with descriptions on same line
                        # or single part number with description
                        if len(matches) == 1:
                            pn = matches[0]
                            # Description is everything after the part number
                            desc = line.replace(pn, '', 1).strip()
                            # If description is empty, check next line
                            if not desc and i + 1 < len(lines):
                                next_l = lines[i+1]
                                if not SAI_PATTERN.search(next_l) and not next_l.startswith('www.'):
                                    desc = next_l
                            products.append({
                                'sai_part_number': pn.upper(),
                                'description': desc,
                                'category': current_category,
                                'brand': 'SAI',
                                'source_page': page_num + 1,
                            })
                        else:
                            # Multiple part numbers on line with potential descriptions
                            # Each part number is likely followed by its description
                            remaining = line
                            for idx, pn in enumerate(matches):
                                pos = remaining.find(pn)
                                if pos >= 0:
                                    after = remaining[pos + len(pn):].strip()
                                    # Find where next part number starts
                                    next_pn_match = SAI_PATTERN.search(after)
                                    if next_pn_match:
                                        desc = after[:next_pn_match.start()].strip()
                                        remaining = after[next_pn_match.start():]
                                    else:
                                        desc = after.strip()
                                        remaining = ''
                                    products.append({
                                        'sai_part_number': pn.upper(),
                                        'description': desc,
                                        'category': current_category,
                                        'brand': 'SAI',
                                        'source_page': page_num + 1,
                                    })
                i += 1

    # Deduplicate by part number (keep last seen - usually more complete description)
    seen = {}
    for p in products:
        pn = p['sai_part_number']
        if pn not in seen or len(p['description']) > len(seen[pn]['description']):
            seen[pn] = p

    products = list(seen.values())

    # Post-process: improve descriptions by doing a second pass
    # Many pages have part numbers on one line, descriptions on next lines
    # Let's do a smarter extraction pass
    print(f"  First pass: {len(products)} SAI products extracted")

    # Second pass: extract from raw text blocks more carefully
    sai_products_v2 = extract_sai_v2()

    # Merge: use v2 descriptions where available (better)
    v2_map = {p['sai_part_number']: p for p in sai_products_v2}

    final_products = []
    seen_pns = set()
    for pn, p in v2_map.items():
        final_products.append(p)
        seen_pns.add(pn)

    # Add any from v1 not in v2
    for pn, p in seen.items():
        if pn not in seen_pns:
            final_products.append(p)

    print(f"  After merge: {len(final_products)} SAI products")
    return final_products


def extract_sai_v2():
    """
    Second-pass SAI extraction using a smarter approach:
    Parse the page as a grid - part numbers and descriptions are often
    laid out in a 2x2 or 4-column grid on each page.
    """
    SAI_PATTERN = re.compile(r'\b(SAI-[\w\d]+(?:-BP)?)\b', re.IGNORECASE)

    # Category detection
    CATEGORIES = [
        "FRONT FENDER", "HEAD LIGHT VISOR", "HEADLIGHT VISOR",
        "SIDE COWL SET", "SIDE COWL", "REAR COWL CENTER PLATE",
        "REAR COWL BACK PLATE", "REAR COWL SET", "REAR COWL",
        "T.P.F.C SET", "T.P.F.C", "ENGINE GUARD", "FUEL TANK COVER",
        "NOSE", "LOWER BODY COVER", "FLOORING UNDER COVER",
        "FRONT BODY COVER", "FOOT TRIM SET", "FOOT TRIM",
        "MUDGUARD", "TAIL PIECE", "SAREE GUARD", "BATTERY COVER",
        "CHAIN COVER", "SILENCER COVER", "SWING ARM COVER",
        "GRAB RAIL", "LEG GUARD", "DASHBOARD", "DASH BOARD",
        "METER VISOR", "HANDLE COVER", "INNER PANEL", "OUTER PANEL",
        "STEP BOARD", "CARRIER PLATE", "REAR CARRIER", "BACK PANEL",
        "WIND SHIELD", "WIND SCREEN", "FAIRING", "VISOR", "COWL",
        "FENDER", "INDICATOR", "TAIL LIGHT", "HEADLIGHT",
    ]
    # Sort longest first for greedy matching
    CATEGORIES.sort(key=lambda x: -len(x))

    products = []
    current_category = "UNKNOWN"

    with pdfplumber.open(SAI_PDF) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if not text:
                continue

            # Detect category on this page
            text_upper = text.upper()
            for cat in CATEGORIES:
                if cat in text_upper:
                    current_category = cat
                    break

            # The page layout typically has:
            # Line with part numbers: SAI-XXX  SAI-YYY  SAI-ZZZ  SAI-WWW
            # Line with descriptions: Desc1  Desc2  Desc3  Desc4
            # (or interleaved: SAI-XXX desc1  SAI-YYY desc2)

            lines = text.split('\n')

            i = 0
            while i < len(lines):
                line = lines[i].strip()
                if not line:
                    i += 1
                    continue

                pn_matches = SAI_PATTERN.findall(line)
                if not pn_matches:
                    i += 1
                    continue

                # Remove page artifacts
                if line.startswith('www.') or line.isdigit():
                    i += 1
                    continue

                # Check if this line is ONLY part numbers
                clean = SAI_PATTERN.sub('', line).strip()

                if not clean:
                    # Pure part number line - descriptions on following lines
                    # Collect description lines until next part-number line
                    desc_lines = []
                    j = i + 1
                    while j < len(lines):
                        nl = lines[j].strip()
                        if not nl:
                            j += 1
                            continue
                        if nl.startswith('www.') or re.match(r'^\d+$', nl):
                            j += 1
                            continue
                        # Stop if this line has part numbers
                        if SAI_PATTERN.search(nl):
                            break
                        desc_lines.append(nl)
                        j += 1

                    # Descriptions are typically split into cols matching the part number count
                    # On a typical page there are 4 part numbers per row, 1 desc line per col
                    n = len(pn_matches)

                    # Try to split description lines into n chunks
                    # Each description line may contain multiple descriptions separated by 2+ spaces
                    all_descs = []
                    for dl in desc_lines:
                        # Split on 2+ spaces to get column descriptions
                        parts = re.split(r'  +', dl)
                        parts = [p.strip() for p in parts if p.strip()]
                        all_descs.extend(parts)

                    # Pair part numbers with descriptions
                    for idx, pn in enumerate(pn_matches):
                        desc = all_descs[idx] if idx < len(all_descs) else ''
                        products.append({
                            'sai_part_number': pn.upper(),
                            'description': desc,
                            'category': current_category,
                            'brand': 'SAI',
                            'source_page': page_num + 1,
                        })

                else:
                    # Mixed line: part numbers with descriptions
                    # Try to extract each part number and its description
                    # by splitting on SAI- boundaries

                    # Find all SAI part number positions
                    for m in SAI_PATTERN.finditer(line):
                        pn = m.group(0).upper()
                        # Get text after this part number until next SAI- or end
                        after_pos = m.end()
                        remaining = line[after_pos:].strip()
                        # Find next part number
                        next_m = SAI_PATTERN.search(remaining)
                        if next_m:
                            desc = remaining[:next_m.start()].strip()
                        else:
                            desc = remaining.strip()

                        # Clean up desc (remove page numbers, website)
                        desc = re.sub(r'\bwww\.\S+', '', desc).strip()
                        desc = re.sub(r'^\d+\s*$', '', desc).strip()

                        products.append({
                            'sai_part_number': pn,
                            'description': desc,
                            'category': current_category,
                            'brand': 'SAI',
                            'source_page': page_num + 1,
                        })

                i += 1

    # Deduplicate - prefer longer descriptions
    seen = {}
    for p in products:
        pn = p['sai_part_number']
        if pn not in seen or len(p['description']) > len(seen[pn]['description']):
            seen[pn] = p

    return list(seen.values())


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 2 – Extract Aerostar products
# ═══════════════════════════════════════════════════════════════════════════════
def extract_aerostar_products():
    """Extract all Aerostar products from the 163-page price list PDF."""
    print("\n═══ TASK 2: Extracting Aerostar products ═══")

    # Part number pattern: AS-XXX, AS-XXXX, ARF-XXX, A82-XX, AMP-XXX, etc.
    AS_PATTERN = re.compile(r'\b(A(?:S|RF|82|MP|RB|RC|F|G|H|T|MP)-[\w\d]+(?:-WP|BP)?)\b', re.IGNORECASE)

    CATEGORIES = [
        "FRONT FENDER", "HEAD LIGHT VISOR", "HEADLIGHT VISOR",
        "SIDE COWL SET", "SIDE COWL", "REAR COWL CENTER PLATE",
        "REAR COWL BACK PLATE", "REAR COWL SET", "REAR COWL",
        "T.P.F.C SET", "T.P.F.C", "ENGINE GUARD", "FUEL TANK COVER",
        "NOSE", "LOWER BODY COVER", "FLOORING UNDER COVER",
        "FRONT BODY COVER", "FOOT TRIM SET", "FOOT TRIM",
        "NEW COLOUR LAUNCH",
    ]
    CATEGORIES.sort(key=lambda x: -len(x))

    BRANDS = ["HERO", "HONDA", "BAJAJ", "TVS", "SUZUKI", "YAMAHA", "KTM",
              "ROYAL ENFIELD", "MAHINDRA", "KAWASAKI"]

    products = []
    current_category = "UNKNOWN"
    current_brand = "UNKNOWN"
    current_vehicle = ""

    with pdfplumber.open(AS_PDF) as pdf:
        total_pages = len(pdf.pages)
        print(f"  Aerostar PDF has {total_pages} pages")

        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if not text:
                continue

            lines = [l.strip() for l in text.split('\n') if l.strip()]

            for line in lines:
                # Skip header line
                if line.startswith("Part No.") and "Description" in line:
                    continue

                # Detect category
                line_upper = line.upper()
                for cat in CATEGORIES:
                    if line_upper == cat or line_upper.startswith(cat):
                        current_category = cat
                        break

                # Detect brand
                for brand in BRANDS:
                    if line_upper == brand:
                        current_brand = brand
                        break

                # Detect vehicle sub-header (FIT FOR ... lines)
                if line_upper.startswith("FIT FOR") or line_upper.startswith("FOR "):
                    current_vehicle = line.strip()
                    continue

                # Try to parse a product line
                # Format: AS-XXXX Description MRP StdPkg [MasterPkg]
                m = AS_PATTERN.match(line)
                if m:
                    part_no = m.group(0).upper()
                    remainder = line[m.end():].strip()

                    # Parse: description, MRP, std_packing
                    # MRP is a number like 384.00 or 1234.00
                    mrp_match = re.search(r'(\d+\.?\d*)\s+(\d+\s+(?:PCS|SET|PC|SETS))', remainder)
                    if mrp_match:
                        description = remainder[:mrp_match.start()].strip()
                        mrp = mrp_match.group(1)
                        std_packing = mrp_match.group(2).strip()
                    else:
                        # Try just MRP at end
                        mrp_only = re.search(r'(\d{3,5}\.?\d*)\s*$', remainder)
                        if mrp_only:
                            description = remainder[:mrp_only.start()].strip()
                            mrp = mrp_only.group(1)
                            std_packing = ""
                        else:
                            description = remainder
                            mrp = ""
                            std_packing = ""

                    # Extract colour from description
                    colour = extract_colour(description)

                    products.append({
                        'as_part_number': part_no,
                        'description': description,
                        'category': current_category,
                        'brand': current_brand,
                        'vehicle': current_vehicle,
                        'colour': colour,
                        'mrp': mrp,
                        'std_packing': std_packing,
                        'source_page': page_num + 1,
                    })

    # Deduplicate by AS part number
    seen = {}
    for p in products:
        pn = p['as_part_number']
        if pn not in seen or len(p['description']) > len(seen[pn]['description']):
            seen[pn] = p

    products = list(seen.values())
    print(f"  Extracted {len(products)} Aerostar products")
    return products


def extract_colour(description):
    """Extract colour from a product description."""
    COLOURS = [
        "Black", "Red", "Blue", "Silver", "Grey", "Green", "White", "Maroon",
        "Navy Blue", "Yellow", "Orange", "Purple", "Gold", "Brown", "Pink",
        "Ivory White", "Sports Red", "Candy Red", "Matte Grey", "Mat Grey",
        "Matte Black", "Metallic Black", "Matte Blue", "V Blue", "CB Red",
        "Parrot Green", "Turquoise Blue", "Selene Silver", "Matt Selene Silver",
        "Dark Blue", "Light Blue", "Spartan Red", "Siren Blue", "Ice Green",
        "Ground Grey", "Matte Brown", "Matte Titanium Grey", "Starlight Blue",
        "Metallic Blue", "Stellar Blue",
    ]
    COLOURS.sort(key=lambda x: -len(x))

    desc_upper = description.upper()
    for colour in COLOURS:
        if colour.upper() in desc_upper:
            return colour
    return ""


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 3 – Match Aerostar ↔ SAI
# ═══════════════════════════════════════════════════════════════════════════════
def match_products(as_products, sai_products):
    """Match Aerostar products to SAI equivalents."""
    print("\n═══ TASK 3: Matching Aerostar ↔ SAI products ═══")

    # Build SAI lookup maps
    sai_by_pn = {}  # base number -> list of SAI products
    for p in sai_products:
        # Extract numeric part: SAI-650A -> 650, SAI-02 -> 02
        base_m = re.search(r'SAI-(\d+)', p['sai_part_number'], re.IGNORECASE)
        if base_m:
            base_num = base_m.group(1)
            sai_by_pn.setdefault(base_num, []).append(p)

    sai_by_full = {p['sai_part_number']: p for p in sai_products}

    matched = []
    unmatched_as = []

    for asp in as_products:
        pn = asp['as_part_number']  # e.g. AS-63, AS-462A

        # Extract number portion from AS part number
        # AS-63 -> 63, AS-462A -> 462, AS-63A -> 63
        num_m = re.search(r'(?:AS|ARF|A82|AMP)-(\d+)', pn, re.IGNORECASE)
        suffix_m = re.search(r'(?:AS|ARF|A82|AMP)-\d+([A-Z](?:[A-Z0-9]*)?)?(?:-WP|-BP)?$', pn, re.IGNORECASE)

        matched_sai = None
        confidence = 'none'

        if num_m:
            base_num = num_m.group(1)
            # Get suffix (A, B, C, etc.)
            suffix = suffix_m.group(1) if suffix_m and suffix_m.group(1) else ''

            # Try exact match: AS-63A -> SAI-63A
            # Remove -WP suffix as SAI uses -BP
            clean_pn = pn.replace('-WP', '')
            sai_equiv_exact = clean_pn.replace('AS-', 'SAI-').replace('ARF-', 'SAI-').replace('A82-', 'SAI-').replace('AMP-', 'SAI-')

            if sai_equiv_exact in sai_by_full:
                matched_sai = sai_by_full[sai_equiv_exact]
                confidence = 'exact_pattern'
            elif '-WP' in pn:
                # WP = Wholesale Packing, same as BP in SAI
                sai_bp = sai_equiv_exact + '-BP'
                # Also try without -BP
                if sai_bp in sai_by_full:
                    matched_sai = sai_by_full[sai_bp]
                    confidence = 'exact_pattern_bp'
                elif sai_equiv_exact in sai_by_full:
                    matched_sai = sai_by_full[sai_equiv_exact]
                    confidence = 'exact_pattern'

            # Try base number match with description similarity
            if not matched_sai and base_num in sai_by_pn:
                candidates = sai_by_pn[base_num]
                # Find best description match
                as_desc = asp['description'].lower()
                best_ratio = 0.0
                best_cand = None
                for cand in candidates:
                    sai_desc = cand['description'].lower()
                    ratio = SequenceMatcher(None, as_desc, sai_desc).ratio()
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_cand = cand

                if best_ratio > 0.5:
                    matched_sai = best_cand
                    confidence = f'description_sim_{best_ratio:.2f}'
                elif best_ratio > 0.3:
                    matched_sai = best_cand
                    confidence = f'low_sim_{best_ratio:.2f}'

        entry = {**asp}
        if matched_sai:
            entry['sai_part_number'] = matched_sai['sai_part_number']
            entry['matched'] = 'yes'
            entry['match_confidence'] = confidence
            matched.append(entry)
        else:
            entry['sai_part_number'] = ''
            entry['matched'] = 'no'
            entry['match_confidence'] = 'none'
            unmatched_as.append(entry)

    print(f"  Matched: {len(matched)} | Unmatched AS: {len(unmatched_as)}")
    return matched + unmatched_as, matched, unmatched_as


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 4 – Re-extract images from SAI PDF
# ═══════════════════════════════════════════════════════════════════════════════
def extract_images_from_sai():
    """Re-extract images from SAI PDF using PyMuPDF with improved spatial matching."""
    print("\n═══ TASK 4: Re-extracting images from SAI PDF ═══")

    SAI_PATTERN = re.compile(r'\b(SAI-[\w\d]+(?:-BP)?)\b', re.IGNORECASE)

    saved = 0
    skipped_small = 0
    skipped_large = 0
    no_label_found = 0
    already_exists = 0

    doc = fitz.open(SAI_PDF)
    print(f"  Processing {len(doc)} pages...")

    for page_num in range(len(doc)):
        page = doc[page_num]

        # Get all text blocks with positions
        blocks = page.get_text("blocks")
        # blocks: (x0, y0, x1, y1, text, block_no, block_type)
        # block_type 0 = text, 1 = image

        text_blocks = [(b[0], b[1], b[2], b[3], b[4].strip())
                       for b in blocks
                       if b[6] == 0 and b[4].strip()]

        # Get all SAI part numbers mentioned on this page with positions
        pn_positions = []
        for (x0, y0, x1, y1, txt) in text_blocks:
            for m in SAI_PATTERN.finditer(txt):
                pn_positions.append({
                    'pn': m.group(0).upper(),
                    'x0': x0, 'y0': y0, 'x1': x1, 'y1': y1,
                    'cx': (x0 + x1) / 2,
                    'cy': (y0 + y1) / 2,
                })

        # Get all images on this page
        img_list = page.get_images(full=True)

        if not img_list:
            continue

        for img_info in img_list:
            xref = img_info[0]

            try:
                base_img = doc.extract_image(xref)
                img_bytes = base_img["image"]
                img_ext = base_img["ext"]
                img_w = base_img["width"]
                img_h = base_img["height"]
            except Exception:
                continue

            # Filter: skip tiny (icons) and huge (backgrounds/banners)
            if img_w < 40 or img_h < 40:
                skipped_small += 1
                continue
            if img_w > 500 and img_h > 200:
                skipped_large += 1
                continue

            # Get image position on page using get_image_rects
            try:
                img_rects = page.get_image_rects(xref)
            except Exception:
                img_rects = []

            if not img_rects:
                # Try to find image rect from image_info
                # Fall back to searching all text on page
                img_cx = page.rect.width / 2
                img_cy = page.rect.height / 2
                img_x0, img_y0 = 0, 0
                img_x1, img_y1 = page.rect.width, page.rect.height
            else:
                rect = img_rects[0]
                img_x0, img_y0, img_x1, img_y1 = rect.x0, rect.y0, rect.x1, rect.y1
                img_cx = (img_x0 + img_x1) / 2
                img_cy = (img_y0 + img_y1) / 2

            # Find nearest SAI part number
            best_pn = None
            best_dist = float('inf')

            for pn_info in pn_positions:
                # Check horizontal overlap: image and text should overlap horizontally
                # (i.e., text is above/below the image in same column)
                horiz_overlap = not (pn_info['x1'] < img_x0 - 20 or pn_info['x0'] > img_x1 + 20)

                # Distance: primarily vertical (above or below within 150px)
                vert_dist = min(abs(pn_info['y1'] - img_y0),  # text above image
                                abs(pn_info['y0'] - img_y1))  # text below image

                # Also consider center-to-center distance
                center_dist = ((pn_info['cx'] - img_cx)**2 + (pn_info['cy'] - img_cy)**2) ** 0.5

                # Score: prefer text close to image, with horizontal overlap
                if horiz_overlap:
                    score = vert_dist
                else:
                    score = center_dist * 2  # penalize non-overlapping

                if score < best_dist and center_dist < 200:
                    best_dist = score
                    best_pn = pn_info['pn']

            if not best_pn:
                # Try full page text search - just find ALL SAI numbers and pick nearest by position
                full_text = page.get_text()
                all_pns = SAI_PATTERN.findall(full_text)
                if all_pns:
                    # Use the first one as fallback (will be marked for review)
                    # But only if there's one unique part number on the page
                    unique_pns = list(set(all_pns))
                    if len(unique_pns) == 1:
                        best_pn = unique_pns[0].upper()
                        best_dist = 999

                if not best_pn:
                    no_label_found += 1
                    continue

            # Save the image
            filename = f"{best_pn}.{img_ext}"
            filepath = IMAGES_DIR / filename

            # Also try .jpeg extension
            if img_ext == 'jpeg' or img_ext == 'jpg':
                filepath_jpeg = IMAGES_DIR / f"{best_pn}.jpeg"
                filepath_png = IMAGES_DIR / f"{best_pn}.png"
            else:
                filepath_jpeg = IMAGES_DIR / f"{best_pn}.jpeg"
                filepath_png = IMAGES_DIR / f"{best_pn}.png"

            # Check if already exists
            if filepath_jpeg.exists() or filepath_png.exists() or filepath.exists():
                already_exists += 1
                continue

            try:
                with open(filepath, 'wb') as f:
                    f.write(img_bytes)
                saved += 1
            except Exception as e:
                print(f"    Error saving {filepath}: {e}")

    doc.close()

    total_images = len(list(IMAGES_DIR.glob('*')))
    print(f"  Saved: {saved} new images")
    print(f"  Already existed: {already_exists}")
    print(f"  Skipped (too small): {skipped_small}")
    print(f"  Skipped (too large): {skipped_large}")
    print(f"  No label found: {no_label_found}")
    print(f"  Total images in directory: {total_images}")
    return total_images


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 5 – Rebuild product_catalog.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def has_image(sai_part_number):
    """Check if an image exists for this SAI part number."""
    if not sai_part_number:
        return False
    for ext in ['jpeg', 'jpg', 'png', 'webp']:
        if (IMAGES_DIR / f"{sai_part_number}.{ext}").exists():
            return True
    return False


def rebuild_catalog(all_products, sai_products, matched_pairs, unmatched_as):
    """Rebuild the product_catalog.xlsx with all sheets."""
    print("\n═══ TASK 5: Rebuilding product_catalog.xlsx ═══")

    # Build set of SAI part numbers that are already matched to Aerostar
    matched_sai_pns = set()
    for p in all_products:
        if p.get('matched') == 'yes' and p.get('sai_part_number'):
            matched_sai_pns.add(p['sai_part_number'])

    # SAI-only products (not matched to any Aerostar product)
    matched_as_pns = set()
    for p in all_products:
        if p.get('matched') == 'yes':
            matched_as_pns.add(p.get('as_part_number', ''))

    sai_only = [p for p in sai_products if p['sai_part_number'] not in matched_sai_pns]
    print(f"  SAI-only (unmatched) products: {len(sai_only)}")

    # Create workbook
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Products ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Products"

    headers = [
        "Category", "Brand", "Vehicle", "AS Part No", "SAI Part No",
        "Description", "Colour", "MRP (₹)", "Std. Packing", "Has Image", "Needs Review"
    ]
    ws1.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    missing_images = []

    for p in all_products:
        sai_pn = p.get('sai_part_number', '')
        img_exists = has_image(sai_pn) if sai_pn else False

        # Needs review: low confidence match, unknown category, or missing image
        confidence = p.get('match_confidence', '')
        needs_review = (
            p.get('matched') == 'no' or
            'low_sim' in confidence or
            p.get('category', '') in ['UNKNOWN', ''] or
            not img_exists
        )

        row = [
            p.get('category', ''),
            p.get('brand', ''),
            p.get('vehicle', ''),
            p.get('as_part_number', ''),
            sai_pn,
            p.get('description', ''),
            p.get('colour', ''),
            p.get('mrp', ''),
            p.get('std_packing', ''),
            'YES' if img_exists else 'NO',
            'YES' if needs_review else 'NO',
        ]
        ws1.append(row)

        if not img_exists:
            missing_images.append({
                'as_part_number': p.get('as_part_number', ''),
                'sai_part_number': sai_pn,
                'description': p.get('description', ''),
                'category': p.get('category', ''),
            })

    # Auto-width columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 2: SAI Only ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("SAI Only")
    sai_headers = ["Category", "SAI Part No", "Description", "Has Image"]
    ws2.append(sai_headers)

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for p in sai_only:
        sai_pn = p['sai_part_number']
        img_exists = has_image(sai_pn)
        ws2.append([
            p.get('category', ''),
            sai_pn,
            p.get('description', ''),
            'YES' if img_exists else 'NO',
        ])
        if not img_exists:
            missing_images.append({
                'as_part_number': '',
                'sai_part_number': sai_pn,
                'description': p.get('description', ''),
                'category': p.get('category', ''),
            })

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 3: Missing Images ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Missing Images")
    mi_headers = ["AS Part No", "SAI Part No", "Description", "Category"]
    ws3.append(mi_headers)

    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Deduplicate missing images
    seen_mi = set()
    for p in missing_images:
        key = p.get('sai_part_number') or p.get('as_part_number')
        if key and key not in seen_mi:
            seen_mi.add(key)
            ws3.append([
                p.get('as_part_number', ''),
                p.get('sai_part_number', ''),
                p.get('description', ''),
                p.get('category', ''),
            ])

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(CATALOG_XL)
    print(f"  Saved catalog: {CATALOG_XL}")
    print(f"  Sheet 'All Products': {len(all_products)} rows")
    print(f"  Sheet 'SAI Only': {len(sai_only)} rows")
    print(f"  Sheet 'Missing Images': {len(seen_mi)} rows")

    return sai_only, missing_images


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 6 – Update aerostar_products.json
# ═══════════════════════════════════════════════════════════════════════════════
def update_json(all_products):
    """Write complete product list to aerostar_products.json."""
    print("\n═══ TASK 6: Updating aerostar_products.json ═══")

    # Only include Aerostar products (those with AS part numbers)
    as_products = [p for p in all_products if p.get('as_part_number')]

    json_data = []
    for p in as_products:
        sai_pn = p.get('sai_part_number', '')
        json_data.append({
            'part_number': p.get('as_part_number', ''),
            'sai_part_number': sai_pn,
            'description': p.get('description', ''),
            'category': p.get('category', ''),
            'brand': p.get('brand', ''),
            'vehicle': p.get('vehicle', ''),
            'colour': p.get('colour', ''),
            'mrp': p.get('mrp', ''),
            'std_packing': p.get('std_packing', ''),
            'matched': p.get('matched', 'no'),
            'match_confidence': p.get('match_confidence', ''),
            'has_image': has_image(sai_pn) if sai_pn else False,
            'image_path': f"static/images/{sai_pn}.jpeg" if (sai_pn and has_image(sai_pn)) else '',
        })

    with open(JSON_OUT, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)

    print(f"  Written {len(json_data)} products to {JSON_OUT}")
    return json_data


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 70)
    print("RUPANI AUTOMOBILES – COMPREHENSIVE CATALOG AUDIT & REBUILD")
    print("=" * 70)

    # Task 1
    sai_products = extract_sai_products()

    # Task 2
    as_products = extract_aerostar_products()

    # Task 3
    all_products, matched, unmatched_as = match_products(as_products, sai_products)

    # Task 4
    total_images = extract_images_from_sai()

    # Task 5
    sai_only, missing_images = rebuild_catalog(all_products, sai_products, matched, unmatched_as)

    # Task 6
    json_data = update_json(all_products)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("AUDIT SUMMARY")
    print("=" * 70)

    products_with_images = sum(1 for p in all_products
                                if has_image(p.get('sai_part_number', '')))
    products_without_images = len(all_products) - products_with_images

    # Category discovery
    existing_categories = set()
    new_categories = set()
    known_cats = {
        "FRONT FENDER", "HEAD LIGHT VISOR", "SIDE COWL SET", "REAR COWL SET",
        "REAR COWL BACK PLATE", "REAR COWL CENTER PLATE", "T.P.F.C SET",
        "ENGINE GUARD", "FUEL TANK COVER", "NOSE", "LOWER BODY COVER",
        "FRONT BODY COVER",
    }
    for p in all_products:
        cat = p.get('category', 'UNKNOWN')
        if cat in known_cats:
            existing_categories.add(cat)
        elif cat != 'UNKNOWN':
            new_categories.add(cat)

    print(f"\n1. Total Aerostar products found in PDF:  {len(as_products):,}")
    print(f"2. Total SAI products found in PDF:       {len(sai_products):,}")
    print(f"3. Successfully matched pairs:             {len(matched):,}")
    print(f"4. Unmatched Aerostar products:            {len(unmatched_as):,}")
    print(f"5. Total images in static/images/:         {total_images:,}")
    print(f"6. Products with images:                   {products_with_images:,}")
    print(f"   Products WITHOUT images:                {products_without_images:,}")

    if new_categories:
        print(f"\n7. New categories discovered:")
        for cat in sorted(new_categories):
            print(f"   - {cat}")
    else:
        print(f"\n7. No new categories discovered beyond known set")

    print(f"\nFiles written:")
    print(f"   {CATALOG_XL}")
    print(f"   {JSON_OUT}")
    print("=" * 70)


if __name__ == '__main__':
    main()
