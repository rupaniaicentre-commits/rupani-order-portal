from flask import Flask, render_template, jsonify, request, send_file
import json
import os
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
import time
import sqlite3
from datetime import datetime
import difflib

try:
    import gspread
    from google.oauth2.service_account import Credentials as SACredentials
    _GSPREAD_AVAILABLE = True
except ImportError:
    _GSPREAD_AVAILABLE = False

import urllib.request
import urllib.parse
import threading

# ── Surepass RC config ────────────────────────────────────────────
# NOTE: Surepass domain is .io (NOT .app). Match token type to domain:
#   Sandbox:    https://sandbox.surepass.io/api/v1/rc/rc-full
#   Production: https://kyc-api.surepass.io/api/v1/rc/rc-full
SUREPASS_TOKEN   = os.environ.get('SUREPASS_TOKEN', '')
SUREPASS_RC_URL  = os.environ.get(
    'SUREPASS_RC_URL',
    'https://kyc-api.surepass.io/api/v1/rc/rc-full'   # production (.io)
)

# ── Green API (WhatsApp) config ───────────────────────────────────
WA_INSTANCE  = os.environ.get('WA_INSTANCE', '7107619441')
WA_TOKEN     = os.environ.get('WA_TOKEN', 'ff2ce0fab0154d8a94fd5153423feb4ab9a76b0a5b5047cf8f')
WA_GROUP_ID  = os.environ.get('WA_GROUP_ID', '120363425235648966@g.us')  # Fiber order grp
WA_BASE      = f'https://api.green-api.com/waInstance{WA_INSTANCE}'

app = Flask(__name__)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0   # disable static file caching
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = BASE_DIR  # data files now live alongside app.py

_products_cache = None
_catalog_mtime  = None  # tracks product_catalog.xlsx modification time

def _load_from_excel(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['All Products']
    products, seen = [], set()
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn = str(row[col['AS Part No']] or '').strip()
        if not pn or pn in seen:
            continue
        seen.add(pn)
        mrp_raw = row[col.get('MRP (₹)', -1)] if 'MRP (₹)' in col else None
        try:
            mrp = float(mrp_raw) if mrp_raw not in (None, '', '-') else None
        except (ValueError, TypeError):
            mrp = None
        products.append({
            'id':             pn,
            'as_part_number': pn,
            'sai_part_number': str(row[col.get('SAI Part No', -1)] or '').strip(),
            'description':    str(row[col.get('Description', -1)] or '').strip(),
            'category':       str(row[col.get('Category', -1)] or '').strip(),
            'brand':          str(row[col.get('Brand', -1)] or '').strip(),
            'vehicle':        str(row[col.get('Vehicle', -1)] or '').strip(),
            'colour':         str(row[col.get('Colour', -1)] or 'N/A').strip(),
            'mrp':            mrp,
            'std_packing':    str(row[col.get('Std. Packing', -1)] or '').strip() if 'Std. Packing' in col else '',
        })
    wb.close()
    return products

def load_products():
    global _products_cache, _catalog_mtime

    catalog_file = os.path.join(BASE_DIR, 'product_catalog.xlsx')
    mapping_file = os.path.join(PARENT_DIR, 'product_mapping.json')
    aerostar_file = os.path.join(PARENT_DIR, 'aerostar_products.json')

    # If catalog Excel exists, use it and reload when file changes
    if os.path.exists(catalog_file):
        mtime = os.path.getmtime(catalog_file)
        if _products_cache is not None and mtime == _catalog_mtime:
            return _products_cache
        try:
            products = _load_from_excel(catalog_file)
            _products_cache = products
            _catalog_mtime  = mtime
            return products
        except Exception as e:
            print(f'[catalog] Failed to load Excel: {e}')

    if _products_cache is not None:
        return _products_cache

    products = []

    if os.path.exists(mapping_file):
        with open(mapping_file) as f:
            mapping = json.load(f)
        seen = set()
        for m in mapping:
            key = m.get('as_part_number', '') or m.get('sai_part_number', '')
            if not key or key in seen:
                continue
            seen.add(key)
            products.append({
                'id': key,
                'as_part_number': m.get('as_part_number', ''),
                'sai_part_number': m.get('sai_part_number', ''),
                'description': m.get('description', ''),
                'category': m.get('category', ''),
                'brand': m.get('brand', ''),
                'vehicle': (m.get('vehicle') or '').strip(),
                'colour': m.get('colour', 'N/A'),
                'mrp': m.get('mrp'),
            })
    elif os.path.exists(aerostar_file):
        with open(aerostar_file) as f:
            raw = json.load(f)
        seen = set()
        for p in raw:
            pn = p.get('part_number', '')
            if pn in seen:
                continue
            seen.add(pn)
            products.append({
                'id': pn,
                'as_part_number': pn,
                'sai_part_number': pn.replace('AS-', 'SAI-'),
                'description': p.get('description', ''),
                'category': p.get('category', ''),
                'brand': p.get('brand', ''),
                'vehicle': (p.get('vehicle') or '').strip(),
                'colour': p.get('colour', 'N/A'),
                'mrp': p.get('mrp'),
            })

    _products_cache = products
    return products


def load_config():
    cfg = os.path.join(BASE_DIR, 'portal_config.json')
    if os.path.exists(cfg):
        with open(cfg) as f:
            data = json.load(f)
    else:
        data = {
            'order_email': 'harshrupani@rupaniautomobiles.com',
            'smtp_host': 'smtp.gmail.com',
            'smtp_port': 587,
            'smtp_user': '',
            'smtp_pass': ''
        }
    # env vars always override file (used in cloud deployments like Railway)
    if os.environ.get('SMTP_USER'):
        data['smtp_user'] = os.environ['SMTP_USER'].strip()
    if os.environ.get('SMTP_PASS'):
        data['smtp_pass'] = os.environ['SMTP_PASS']  # keep spaces – Gmail App Password uses them
    if os.environ.get('ORDER_EMAIL'):
        data['order_email'] = os.environ['ORDER_EMAIL'].strip()
    if os.environ.get('GSHEET_ID'):
        data['gsheet_id'] = os.environ['GSHEET_ID'].strip()
    if os.environ.get('SMTP_HOST'):
        data['smtp_host'] = os.environ['SMTP_HOST'].strip()
    if os.environ.get('SMTP_PORT'):
        data['smtp_port'] = int(os.environ['SMTP_PORT'].strip())
    return data


def save_config(data):
    cfg = os.path.join(BASE_DIR, 'portal_config.json')
    with open(cfg, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/')
def landing():
    """Portal entry — choose Aerostar (fibre) or Honda."""
    return render_template('choose.html')


@app.route('/aerostar')
def index():
    return render_template('index.html')


@app.route('/honda')
def honda():
    return render_template('honda.html')


# ── Honda parts data (Parent = HONDA 18%, sorted by vehicle) ──────────
_honda_cache = None

def load_honda():
    global _honda_cache
    if _honda_cache is None:
        with open(os.path.join(BASE_DIR, 'honda_parts.json'), encoding='utf-8') as f:
            _honda_cache = json.load(f)
    return _honda_cache


@app.route('/api/honda/data')
def get_honda_data():
    return jsonify(load_honda())


@app.route('/api/honda/new')
def honda_new_public():
    """Newly received Honda parts (for the customer 'New Arrivals' section)."""
    sync_honda_registry()
    try:
        conn = _orders_conn()
        rows = conn.execute("SELECT part_no,batch,first_seen FROM honda_registry "
                            "WHERE batch != 'baseline' ORDER BY first_seen DESC, part_no").fetchall()
        conn.close()
    except Exception as e:
        print(f"[HONDA NEW PUBLIC ERROR] {e}", flush=True)
        return jsonify({'total': 0, 'batches': []})
    batches, order = {}, []
    for pn, batch, fs in rows:
        if batch not in batches:
            batches[batch] = []; order.append(batch)
        batches[batch].append(pn)
    return jsonify({'total': len(rows),
                    'batches': [{'batch': b, 'part_nos': batches[b]} for b in order]})


@app.route('/api/honda/feedback', methods=['POST'])
def honda_feedback():
    """Customer fitment feedback: wrong fitment / fits more vehicles."""
    data = request.json or {}
    part_no = (data.get('part_no') or '').strip()
    comment = (data.get('comment') or '').strip()
    if not part_no or not comment:
        return jsonify({'success': False, 'error': 'Part number and comment required'})

    entry = {
        'ts': datetime.now().isoformat(timespec='seconds'),
        'part_no': part_no,
        'part_name': (data.get('name') or '').strip(),
        'current_fitment': data.get('vehicles') or [],
        'comment': comment,
        'firm': (data.get('firm') or '').strip(),
        'contact': (data.get('contact') or '').strip(),
    }
    try:
        with open(os.path.join(BASE_DIR, 'honda_feedback.jsonl'), 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry, ensure_ascii=False) + '\n')
    except Exception as e:
        print(f"[FEEDBACK FILE ERROR] {e}", flush=True)

    def _notify_feedback(e):
        try:
            msg = ("📝 *Honda fitment feedback*\n"
                   f"🏪 {e['firm']}  📱 {e['contact']}\n"
                   f"🔩 {e['part_no']} — {e['part_name']}\n"
                   f"🚗 Currently under: {', '.join(e['current_fitment']) or '—'}\n"
                   f"💬 {e['comment']}")
            _wa_post("sendMessage", {"chatId": WA_GROUP_ID, "message": msg})
        except Exception as ex:
            print(f"[FEEDBACK WA ERROR] {ex}", flush=True)
    threading.Thread(target=_notify_feedback, args=(entry,), daemon=True).start()

    return jsonify({'success': True})


# ── server-side order history (cross-device, keyed by mobile) ─────────
# Uses SQLite. Set ORDERS_DB to a path on a Railway *volume* (e.g.
# /data/orders.db) so history survives redeploys; otherwise it falls back
# to a file in the app dir (works, but resets on each deploy).
def _orders_db_path():
    if os.environ.get('ORDERS_DB'):
        return os.environ['ORDERS_DB']
    # Railway exposes the attached volume's mount path here (any mount point)
    vol = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH')
    if vol and os.path.isdir(vol):
        return os.path.join(vol, 'orders.db')
    if os.path.isdir('/data'):
        return os.path.join('/data', 'orders.db')
    return os.path.join(BASE_DIR, 'orders.db')

ORDERS_DB = _orders_db_path()
_ON_VOLUME = bool(os.environ.get('ORDERS_DB') or os.environ.get('RAILWAY_VOLUME_MOUNT_PATH') or os.path.isdir('/data'))

def _orders_conn():
    conn = sqlite3.connect(ORDERS_DB, timeout=5)
    conn.execute('''CREATE TABLE IF NOT EXISTS orders(
        oid TEXT PRIMARY KEY, ts INTEGER, portal TEXT, firm TEXT,
        contact TEXT, total_qty INTEGER, total_amt REAL, items TEXT)''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_contact ON orders(contact)')
    cols = [r[1] for r in conn.execute('PRAGMA table_info(orders)').fetchall()]
    if 'status' not in cols:
        conn.execute("ALTER TABLE orders ADD COLUMN status TEXT DEFAULT 'pending'")
    if 'updated' not in cols:
        conn.execute("ALTER TABLE orders ADD COLUMN updated INTEGER DEFAULT 0")
    # admin sessions live in the DB so tokens work across all gunicorn workers
    conn.execute('''CREATE TABLE IF NOT EXISTS sessions(
        token TEXT PRIMARY KEY, "user" TEXT, role TEXT, scope TEXT, exp REAL)''')
    # registry of Honda parts + when each first appeared (to detect new stock)
    conn.execute('''CREATE TABLE IF NOT EXISTS honda_registry(
        part_no TEXT PRIMARY KEY, name TEXT, price REAL, vehicles TEXT,
        first_seen REAL, batch TEXT)''')
    return conn

def _order_status(items):
    total = sum(int(i.get('qty', 0) or 0) for i in items)
    done  = sum(min(int(i.get('disp', 0) or 0), int(i.get('qty', 0) or 0)) for i in items)
    if total and done >= total:
        return 'dispatched'
    return 'partial' if done > 0 else 'pending'

def _num(v):
    s = str(v).strip()
    if s in ('', 'None'):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def save_order_db(oid, portal, firm, contact, items):
    contact_d = re.sub(r'\D', '', contact or '')
    if not contact_d or not items:
        return
    norm = [{'part_no': i.get('as_part_number', ''), 'name': i.get('description', ''),
             'price': _num(i.get('mrp')), 'qty': int(i.get('qty', 0) or 0), 'disp': 0} for i in items]
    tq = sum(n['qty'] for n in norm)
    ta = sum((n['price'] or 0) * n['qty'] for n in norm)
    if not oid:
        oid = f"{int(time.time()*1000)}-{contact_d[-4:]}"
    try:
        conn = _orders_conn()
        conn.execute('INSERT OR IGNORE INTO orders(oid,ts,portal,firm,contact,total_qty,total_amt,items,status,updated) '
                     'VALUES(?,?,?,?,?,?,?,?,?,?)',
                     (oid, int(time.time()*1000), portal, firm, contact_d, tq, ta,
                      json.dumps(norm, ensure_ascii=False), 'pending', 0))
        conn.commit(); conn.close()
    except Exception as e:
        print(f"[ORDER DB WRITE ERROR] {e}", flush=True)

@app.route('/api/orders')
def get_orders():
    contact_d = re.sub(r'\D', '', request.args.get('contact', '') or '')
    if not contact_d:
        return jsonify([])
    try:
        conn = _orders_conn()
        rows = conn.execute(
            'SELECT oid,ts,portal,total_qty,total_amt,items,status FROM orders '
            'WHERE contact=? ORDER BY ts DESC LIMIT 50', (contact_d,)).fetchall()
        conn.close()
        return jsonify([{'oid': r[0], 'ts': r[1], 'portal': r[2], 'totalQty': r[3],
                         'totalAmt': r[4], 'items': json.loads(r[5]), 'status': r[6] or 'pending'}
                        for r in rows])
    except Exception as e:
        print(f"[ORDER DB READ ERROR] {e}", flush=True)
        return jsonify([])


# ── internal dashboard: auth + fulfillment + analytics ────────────────
ADMIN_USERS = {
    'HARSH': {'password': os.environ.get('ADMIN_PW_HARSH', 'RUPANI10'),     'role': 'admin',   'scope': 'all'},
    'ABHAY': {'password': os.environ.get('ADMIN_PW_ABHAY', 'RUPANIABHAY123'), 'role': 'manager', 'scope': 'honda'},
    'DEVA':  {'password': os.environ.get('ADMIN_PW_DEVA',  'RUPANIDEVA123'),  'role': 'manager', 'scope': 'aerostar'},
}
def _admin_auth(token):
    if not token:
        return None
    try:
        conn = _orders_conn()
        row = conn.execute('SELECT "user",role,scope,exp FROM sessions WHERE token=?', (token,)).fetchone()
        conn.close()
    except Exception as e:
        print(f"[SESSION READ ERROR] {e}", flush=True)
        return None
    if row and row[3] > time.time():
        return {'user': row[0], 'role': row[1], 'scope': row[2]}
    return None


@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    d = request.json or {}
    u = (d.get('username') or '').strip().upper()
    p = (d.get('password') or '')
    rec = ADMIN_USERS.get(u)
    if not rec or rec['password'] != p:
        return jsonify({'success': False, 'error': 'Invalid username or password'})
    import secrets
    token = secrets.token_urlsafe(24)
    try:
        conn = _orders_conn()
        conn.execute('DELETE FROM sessions WHERE exp < ?', (time.time(),))   # prune expired
        conn.execute('INSERT INTO sessions(token,"user",role,scope,exp) VALUES(?,?,?,?,?)',
                     (token, u, rec['role'], rec['scope'], time.time() + 86400 * 30))
        conn.commit(); conn.close()
    except Exception as e:
        print(f"[SESSION WRITE ERROR] {e}", flush=True)
        return jsonify({'success': False, 'error': 'session store error'})
    return jsonify({'success': True, 'token': token, 'user': u.title(),
                    'role': rec['role'], 'scope': rec['scope']})


# ─────────────── ADMIN-ONLY: Chassis → Model → Parts TEST page ───────────────
@app.route('/vin-test')
def vin_test_page():
    return render_template('vin_test.html')


def _epc_parts_for_model(model_id, limit=400):
    """Fetch EPC parts for a resolved model_id, grouped by section/illustration."""
    dbp = os.path.join(BASE_DIR, 'epc_parts.db')
    if not os.path.exists(dbp):
        return None, 0
    conn = sqlite3.connect(dbp)
    total = conn.execute('SELECT COUNT(*) FROM parts WHERE model_id=?', (str(model_id),)).fetchone()[0]
    rows = conn.execute(
        'SELECT section,illus,pn,descr,qty,ns FROM parts WHERE model_id=? ORDER BY section,illus LIMIT ?',
        (str(model_id), limit)).fetchall()
    conn.close()
    groups = {}
    for section, illus, pn, descr, qty, ns in rows:
        key = f'{section} · {illus}'
        groups.setdefault(key, []).append({'pn': pn, 'desc': descr, 'qty': qty, 'ns': ns})
    return [{'group': k, 'parts': v} for k, v in groups.items()], total


def _surepass_diag(reg_number):
    """Admin-only diagnostic: shows exactly why an RC (VAHAN) lookup failed."""
    d = {'token_set': bool(SUREPASS_TOKEN),
         'url': SUREPASS_RC_URL,
         'url_type': 'PRODUCTION' if 'kyc-api' in SUREPASS_RC_URL else 'SANDBOX'}
    if not SUREPASS_TOKEN:
        d['reason'] = 'SUREPASS_TOKEN Railway env me set nahi hai'
        return d
    try:
        payload = json.dumps({'id_number': reg_number}).encode('utf-8')
        req = urllib.request.Request(
            SUREPASS_RC_URL, data=payload, method='POST',
            headers={'Authorization': f'Bearer {SUREPASS_TOKEN}',
                     'Content-Type': 'application/json', 'Accept': 'application/json',
                     'User-Agent': 'RupaniOrderPortal/1.0'})
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw = json.loads(resp.read().decode())
        d['status_code'] = raw.get('status_code')
        d['success'] = raw.get('success')
        d['message'] = raw.get('message')
        data = raw.get('data') or {}
        d['maker_model'] = data.get('maker_model')
        d['has_chassis'] = bool(data.get('vehicle_chasi_number'))
    except urllib.error.HTTPError as e:
        d['http_error'] = e.code
        try: d['body'] = e.read().decode('utf-8', 'replace')[:300]
        except Exception: pass
    except Exception as e:
        d['error'] = str(e)[:200]
    return d


@app.route('/api/admin/resolve-vin', methods=['POST'])
def admin_resolve_vin():
    auth = _admin_auth((request.json or {}).get('token', ''))
    if not auth or auth.get('role') != 'admin':          # admin-only (HARSH)
        return jsonify({'success': False, 'error': 'unauthorized — admin only'}), 401
    d = request.json or {}
    try:
        import vds_resolver as R
    except Exception as e:
        return jsonify({'success': False, 'error': f'resolver load failed: {e}'})

    pick = (d.get('pick') or '').strip()               # a code chosen after a filter question
    if pick:
        model_id = R._mid(pick)
        parts, total = _epc_parts_for_model(model_id)
        return jsonify({'success': True, 'resolved': True, 'model_code': pick,
                        'model_id': model_id, 'total_parts': total, 'groups': parts})

    # Primary input: vehicle (registration) number -> Surepass RC -> chassis + model
    reg = (d.get('reg_number') or '').strip().upper().replace(' ', '')
    vehicle = None
    chassis = (d.get('chassis') or '').strip() or None
    maker_model = (d.get('maker_model') or '').strip() or None
    mfg_year = int(d['mfg_year']) if str(d.get('mfg_year') or '').strip().isdigit() else None
    norms = (d.get('norms') or '').strip() or None
    if reg:
        info = _fetch_vehicle_info(reg)
        if not info:
            return jsonify({'success': True, 'resolved': False,
                            'error': 'Vehicle number nahi mila (VAHAN se detail nahi aayi). '
                                     'Number sahi hai? Ya API token check karo.',
                            'diag': _surepass_diag(reg)})
        vehicle = {'rc_number': info.get('rc_number'), 'owner': info.get('owner_name'),
                   'model': info.get('maker_model'), 'year': info.get('year'),
                   'colour': info.get('colour'), 'fuel': info.get('fuel_type'),
                   'vehicle_class': info.get('vehicle_class')}
        chassis = info.get('chassis') or chassis
        maker_model = info.get('maker_model') or maker_model
        if not mfg_year and str(info.get('year') or '').isdigit():
            mfg_year = int(info['year'])
        norms = info.get('norms_type') or norms
        if not maker_model and not chassis:
            return jsonify({'success': True, 'resolved': False, 'vehicle': vehicle,
                            'error': 'VAHAN se is gaadi ka model/chassis nahi mila. '
                                     'Number sahi hai kya? (ya yeh Honda nahi hai)'})
        make = (info.get('make') or '').upper()
        if make and 'HONDA' not in make:
            return jsonify({'success': True, 'resolved': False, 'vehicle': vehicle,
                            'error': f'Yeh Honda gaadi nahi hai ({info.get("make")}). Abhi sirf Honda ke parts hain.'})

    res = R.resolve(chassis=chassis, maker_model=maker_model, mfg_year=mfg_year, norms=norms)
    if not res.get('ok'):
        return jsonify({'success': True, 'resolved': False, 'error': res.get('reason'),
                        'vds4': res.get('vds4'), 'year': res.get('year'), 'vehicle': vehicle})
    if res.get('needs_filter'):
        return jsonify({'success': True, 'resolved': False, 'needs_filter': True,
                        'family': res['family'], 'year': res['year'],
                        'candidates': res['candidates'], 'question': res['question'], 'vehicle': vehicle})
    parts, total = _epc_parts_for_model(res['model_id'])
    return jsonify({'success': True, 'resolved': True, 'family': res['family'], 'year': res['year'],
                    'model_code': res['model_code'], 'model_id': res['model_id'],
                    'note': res.get('note'), 'total_parts': total, 'groups': parts, 'vehicle': vehicle})


@app.route('/api/admin/orders')
def admin_orders():
    auth = _admin_auth(request.args.get('token', ''))
    if not auth:
        return jsonify({'success': False, 'error': 'unauthorized'}), 401
    portal = request.args.get('portal', '')
    scope = auth['scope']
    where, args = [], []
    if scope in ('honda', 'aerostar'):
        where.append('portal=?'); args.append(scope)
    elif portal in ('honda', 'aerostar'):
        where.append('portal=?'); args.append(portal)
    q = 'SELECT oid,ts,portal,firm,contact,total_qty,total_amt,items,status FROM orders'
    if where:
        q += ' WHERE ' + ' AND '.join(where)
    q += ' ORDER BY ts DESC LIMIT 2000'
    conn = _orders_conn()
    rows = conn.execute(q, args).fetchall()
    conn.close()
    orders = [{'oid': r[0], 'ts': r[1], 'portal': r[2], 'firm': r[3], 'contact': r[4],
               'totalQty': r[5], 'totalAmt': r[6], 'items': json.loads(r[7]),
               'status': r[8] or 'pending'} for r in rows]
    return jsonify({'success': True, 'role': auth['role'], 'scope': scope, 'orders': orders})


@app.route('/api/admin/dispatch', methods=['POST'])
def admin_dispatch():
    d = request.json or {}
    auth = _admin_auth(d.get('token', ''))
    if not auth:
        return jsonify({'success': False, 'error': 'unauthorized'}), 401
    oid = d.get('oid')
    updates = {u.get('part_no'): u for u in (d.get('items') or [])}
    mark_all = d.get('mark_all')
    conn = _orders_conn()
    row = conn.execute('SELECT portal,items FROM orders WHERE oid=?', (oid,)).fetchone()
    if not row:
        conn.close(); return jsonify({'success': False, 'error': 'not found'})
    portal, items = row[0], json.loads(row[1])
    if auth['scope'] in ('honda', 'aerostar') and auth['scope'] != portal:
        conn.close(); return jsonify({'success': False, 'error': 'forbidden'}), 403
    for it in items:
        if mark_all:
            it['disp'] = it['qty']
        elif it['part_no'] in updates:
            u = updates[it['part_no']]
            it['disp'] = max(0, min(int(u.get('disp', 0) or 0), it['qty']))
            if 'alt' in u:                       # alternate part number actually sent
                it['alt'] = (u.get('alt') or '').strip()
    status = _order_status(items)
    conn.execute('UPDATE orders SET items=?, status=?, updated=? WHERE oid=?',
                 (json.dumps(items, ensure_ascii=False), status, int(time.time()*1000), oid))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'status': status, 'items': items})


_registry_synced = False

def sync_honda_registry():
    """Record any Honda parts not seen before. First run = 'baseline' (not new);
    later runs tag newly-appeared parts with the sync date so the dashboard can
    show 'new items received'."""
    global _registry_synced
    if _registry_synced:
        return
    try:
        parts = load_honda().get('parts', [])
        conn = _orders_conn()
        existing = set(r[0] for r in conn.execute('SELECT part_no FROM honda_registry').fetchall())
        batch = 'baseline' if not existing else datetime.now().strftime('%Y-%m-%d')
        now = time.time()
        added = 0
        for p in parts:
            pn = p.get('part_no')
            if not pn or pn in existing:
                continue
            conn.execute('INSERT OR IGNORE INTO honda_registry VALUES(?,?,?,?,?,?)',
                         (pn, p.get('name', ''), p.get('price'),
                          ' / '.join([v for v in (p.get('vehicles') or []) if v != 'OTHER']),
                          now, batch))
            added += 1
        conn.commit(); conn.close()
        if added:
            print(f"[HONDA REGISTRY] +{added} parts (batch={batch})", flush=True)
        _registry_synced = True
    except Exception as e:
        print(f"[HONDA REGISTRY ERROR] {e}", flush=True)


@app.route('/api/admin/honda-new')
def admin_honda_new():
    auth = _admin_auth(request.args.get('token', ''))
    if not auth:
        return jsonify({'success': False, 'error': 'unauthorized'}), 401
    if auth['scope'] == 'aerostar':
        return jsonify({'success': True, 'batches': [], 'total': 0})
    sync_honda_registry()
    conn = _orders_conn()
    rows = conn.execute("SELECT part_no,name,price,vehicles,first_seen,batch FROM honda_registry "
                        "WHERE batch != 'baseline' ORDER BY first_seen DESC, part_no").fetchall()
    conn.close()
    batches = {}
    order = []
    for pn, name, price, veh, fs, batch in rows:
        if batch not in batches:
            batches[batch] = []
            order.append(batch)
        batches[batch].append({'part_no': pn, 'name': name, 'price': price, 'vehicles': veh})
    out = [{'batch': b, 'count': len(batches[b]), 'items': batches[b]} for b in order]
    return jsonify({'success': True, 'batches': out, 'total': len(rows)})


@app.route('/api/admin/diag')
def admin_diag():
    auth = _admin_auth(request.args.get('token', ''))
    if not auth:
        return jsonify({'success': False, 'error': 'unauthorized'}), 401
    info = {'success': True, 'db_path': ORDERS_DB, 'on_volume': _ON_VOLUME,
            'volume_env': os.environ.get('RAILWAY_VOLUME_MOUNT_PATH') or '',
            'data_dir_exists': os.path.isdir('/data')}
    try:
        conn = _orders_conn()
        info['order_count'] = conn.execute('SELECT count(*) FROM orders').fetchone()[0]
        conn.close()
        info['writable'] = os.access(os.path.dirname(ORDERS_DB) or '.', os.W_OK)
    except Exception as e:
        info['writable'] = False
        info['error'] = str(e)
    return jsonify(info)


@app.route('/api/admin/pending-export')
def admin_pending_export():
    auth = _admin_auth(request.args.get('token', ''))
    if not auth:
        return jsonify({'success': False, 'error': 'unauthorized'}), 401
    scope = auth['scope']
    portal = scope if scope in ('honda', 'aerostar') else (request.args.get('portal') or None)
    conn = _orders_conn()
    rows = conn.execute(
        'SELECT ts,portal,firm,contact,items FROM orders'
        + (' WHERE portal=?' if portal in ('honda', 'aerostar') else '')
        + ' ORDER BY ts',
        ([portal] if portal in ('honda', 'aerostar') else [])).fetchall()
    conn.close()

    agg, lines = {}, []
    for ts, p, firm, contact, items_j in rows:
        for it in json.loads(items_j):
            qn = int(it.get('qty', 0) or 0)
            dp = min(int(it.get('disp', 0) or 0), qn)
            pend = qn - dp
            if pend <= 0:
                continue
            pr = it.get('price') or 0
            a = agg.setdefault(it['part_no'], {'name': it['name'], 'pending': 0, 'val': 0,
                                               'ordered': 0, 'sent': 0, 'portal': p})
            a['pending'] += pend; a['val'] += pr * pend; a['ordered'] += qn; a['sent'] += dp
            lines.append((datetime.fromtimestamp(ts / 1000).strftime('%d-%m-%Y'), p, firm, contact,
                          it['part_no'], it['name'], qn, dp, pend, pr, pr * pend))

    navy = PatternFill('solid', fgColor='1B2A4A')
    hf = Font(color='FFFFFF', bold=True, size=11)
    bold = Font(bold=True)
    thin = Side(style='thin', color='CCCCCC')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    label = (portal or 'All').title()

    # ── Sheet 1: procurement summary (what to send the supplier) ──
    ws = wb.active; ws.title = 'Pending Summary'
    ws.merge_cells('A1:F1')
    ws['A1'] = f'RUPANI AUTOMOBILES — PENDING PROCUREMENT ({label})  |  {datetime.now().strftime("%d-%m-%Y %H:%M")}'
    ws['A1'].font = Font(bold=True, size=13, color='1B2A4A')
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 26
    hdr = ['Part No', 'Description', 'Total Ordered', 'Already Sent', 'Pending Qty', 'Pending Value (₹)']
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(3, c, h); cell.fill = navy; cell.font = hf; cell.border = bd; cell.alignment = ctr
    r = 4
    for pn, a in sorted(agg.items(), key=lambda kv: -kv[1]['pending']):
        for c, v in enumerate([pn, a['name'], a['ordered'], a['sent'], a['pending'], round(a['val'])], 1):
            cell = ws.cell(r, c, v); cell.border = bd
            if c in (3, 4, 5, 6):
                cell.alignment = Alignment(horizontal='center')
        r += 1
    ws.cell(r, 4, 'TOTAL').font = bold
    ws.cell(r, 5, sum(a['pending'] for a in agg.values())).font = bold
    ws.cell(r, 6, round(sum(a['val'] for a in agg.values()))).font = bold
    for col, w in zip('ABCDEF', [16, 46, 14, 13, 12, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

    # ── Sheet 2: detailed pending order lines (internal reference) ──
    ws2 = wb.create_sheet('Pending Order Lines')
    hdr2 = ['Order Date', 'Catalogue', 'Firm', 'Mobile', 'Part No', 'Description',
            'Ordered', 'Sent', 'Pending', 'Rate', 'Pending Value']
    for c, h in enumerate(hdr2, 1):
        cell = ws2.cell(1, c, h); cell.fill = navy; cell.font = hf; cell.border = bd; cell.alignment = ctr
    for i, row in enumerate(lines, 2):
        for c, v in enumerate(row, 1):
            cell = ws2.cell(i, c, (round(v) if isinstance(v, float) else v)); cell.border = bd
    for col, w in zip('ABCDEFGHIJK', [12, 11, 22, 13, 15, 40, 9, 8, 9, 9, 13]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A2'

    fname = f"Pending_{label}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(tempfile.gettempdir(), fname)
    wb.save(path)
    return send_file(path, as_attachment=True, download_name=fname)


@app.route('/api/admin/analytics')
def admin_analytics():
    auth = _admin_auth(request.args.get('token', ''))
    if not auth:
        return jsonify({'success': False, 'error': 'unauthorized'}), 401
    scope = auth['scope']
    portal = scope if scope in ('honda', 'aerostar') else (request.args.get('portal') or None)
    where, args = [], []
    if portal in ('honda', 'aerostar'):
        where.append('portal=?'); args.append(portal)
    q = 'SELECT total_amt,items,status FROM orders'
    if where:
        q += ' WHERE ' + ' AND '.join(where)
    conn = _orders_conn()
    rows = conn.execute(q, args).fetchall()
    conn.close()

    ordered_val = disp_val = pend_val = ordered_qty = disp_qty = 0
    status_ct = {'pending': 0, 'partial': 0, 'dispatched': 0}
    part_agg = {}
    for amt, items_j, status in rows:
        status_ct[status] = status_ct.get(status, 0) + 1
        for it in json.loads(items_j):
            q_ = int(it.get('qty', 0) or 0)
            dp = min(int(it.get('disp', 0) or 0), q_)
            pr = it.get('price') or 0
            ordered_qty += q_; disp_qty += dp
            ordered_val += pr * q_; disp_val += pr * dp; pend_val += pr * (q_ - dp)
            a = part_agg.setdefault(it['part_no'], {'part_no': it['part_no'], 'name': it['name'],
                                                    'ordered': 0, 'disp': 0, 'val': 0})
            a['ordered'] += q_; a['disp'] += dp; a['val'] += pr * q_
    for a in part_agg.values():
        a['pending'] = a['ordered'] - a['disp']
        a['pending_val'] = 0  # filled below
    top_pending = sorted([a for a in part_agg.values() if a['pending'] > 0],
                         key=lambda x: -x['pending'])[:25]
    top_demand = sorted(part_agg.values(), key=lambda x: -x['ordered'])[:25]
    fill = round(100 * disp_qty / ordered_qty) if ordered_qty else 0
    return jsonify({'success': True, 'portal': portal or 'all', 'n_orders': len(rows),
                    'ordered_val': ordered_val, 'disp_val': disp_val, 'pend_val': pend_val,
                    'ordered_qty': ordered_qty, 'disp_qty': disp_qty, 'fill_ratio': fill,
                    'status_counts': status_ct, 'top_pending': top_pending, 'top_demand': top_demand})


@app.route('/api/products')
def get_products():
    return jsonify(load_products())


@app.route('/api/config', methods=['GET'])
def get_config():
    c = load_config()
    return jsonify({k: v for k, v in c.items() if k != 'smtp_pass'})


@app.route('/api/config', methods=['POST'])
def update_config():
    data = request.json or {}
    c = load_config()
    c.update(data)
    save_config(c)
    return jsonify({'success': True})


@app.route('/api/test-email')
def test_email():
    config = load_config()
    smtp_user = config.get('smtp_user', '')
    smtp_pass = config.get('smtp_pass', '')
    order_email = config.get('order_email', '')
    smtp_host = config.get('smtp_host', 'smtp.gmail.com')
    smtp_port = int(config.get('smtp_port', 587))

    print(f"[TEST-EMAIL] smtp_user={repr(smtp_user)}", flush=True)
    print(f"[TEST-EMAIL] smtp_pass_len={len(smtp_pass)} smtp_pass_repr={repr(smtp_pass)}", flush=True)
    print(f"[TEST-EMAIL] order_email={repr(order_email)}", flush=True)
    print(f"[TEST-EMAIL] smtp_host={smtp_host} smtp_port={smtp_port}", flush=True)

    if not smtp_user or not smtp_pass:
        return jsonify({
            'success': False,
            'smtp_user': smtp_user,
            'smtp_pass_len': len(smtp_pass),
            'error': 'Missing SMTP credentials'
        })

    try:
        msg = MIMEText('Test email from Rupani Order Portal (/api/test-email)')
        msg['From'] = smtp_user
        msg['To'] = order_email
        msg['Subject'] = 'Rupani Portal – SMTP Test'
        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, [order_email], msg.as_string())
        return jsonify({'success': True, 'smtp_user': smtp_user, 'smtp_pass_len': len(smtp_pass), 'error': None})
    except Exception as e:
        print(f"[TEST-EMAIL ERROR] {e}", flush=True)
        return jsonify({'success': False, 'smtp_user': smtp_user, 'smtp_pass_len': len(smtp_pass), 'error': str(e)})


@app.route('/api/reload-catalog', methods=['POST'])
def reload_catalog():
    global _products_cache, _catalog_mtime
    _products_cache = None
    _catalog_mtime  = None
    load_products()
    return jsonify({'success': True, 'total': len(_products_cache or [])})


@app.route('/api/checkout', methods=['POST'])
def checkout():
    data = request.json or {}
    firm_name = (data.get('firm_name') or '').strip()
    contact_number = (data.get('contact_number') or '').strip()
    items = data.get('items', [])
    portal = (data.get('portal') or 'aerostar').strip().lower()
    oid = (data.get('oid') or '').strip()

    if not items:
        return jsonify({'success': False, 'error': 'No items in order'})
    if not firm_name:
        return jsonify({'success': False, 'error': 'Firm name required'})

    # persist to cross-device order history (keyed by mobile)
    save_order_db(oid, portal, firm_name, contact_number, items)

    wb = _build_excel(firm_name, contact_number, items, portal)

    fname = f"Order_{firm_name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_path = os.path.join(BASE_DIR, fname)
    wb.save(save_path)

    config = load_config()
    smtp_user = config.get('smtp_user', '')
    smtp_pass = config.get('smtp_pass', '')
    order_email = config.get('order_email', 'harshrupani@rupaniautomobiles.com')

    # ── Fire WhatsApp + GSheet in background, return immediately ─
    def _notify(sp, fn, ct, it):
        try:
            _send_whatsapp(fn, ct, it, sp, fname, portal)
        except Exception as e:
            print(f"[WA ERROR] {type(e).__name__}: {e}", flush=True)
        try:
            _log_to_gsheet(config, fn, ct, it, False)
        except Exception as e:
            print(f"[GSHEET ERROR] {type(e).__name__}: {e}", flush=True)

    threading.Thread(
        target=_notify,
        args=(save_path, firm_name, contact_number, items),
        daemon=True
    ).start()

    return jsonify({'success': True, 'email_sent': False, 'wa_sent': True,
                    'message': '✅ Order placed! Notifying Fiber order grp…', 'download': fname})


@app.route('/download/<path:filename>')
def download_order(filename):
    fp = os.path.join(BASE_DIR, filename)
    if os.path.exists(fp):
        return send_file(fp, as_attachment=True, download_name=filename)
    return 'File not found', 404


def _build_excel(firm_name, contact_number, items, portal='aerostar'):
    is_honda = portal == 'honda'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order'

    navy = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    gold_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    hdr_font = Font(color='FFFFFF', bold=True, size=11)
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    title_font = Font(bold=True, size=14, color='1B2A4A')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = 'RUPANI AUTOMOBILES PVT. LTD. — ORDER SHEET'
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws['A1'].fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    ws.row_dimensions[1].height = 30

    ws.append([])  # row 2

    # Row 3: Firm Name
    ws['A3'] = 'Firm Name'
    ws['A3'].font = bold_font
    ws['A3'].fill = gold_fill
    ws['A3'].border = border
    ws['B3'] = firm_name
    ws['B3'].font = normal_font
    ws['B3'].border = border
    ws.merge_cells('B3:D3')

    # Row 4: Contact
    ws['A4'] = 'Contact Number'
    ws['A4'].font = bold_font
    ws['A4'].fill = gold_fill
    ws['A4'].border = border
    ws['B4'] = contact_number
    ws['B4'].font = normal_font
    ws['B4'].border = border
    ws.merge_cells('B4:D4')

    # Row 5: Date
    ws['A5'] = 'Order Date'
    ws['A5'].font = bold_font
    ws['A5'].fill = gold_fill
    ws['A5'].border = border
    ws['B5'] = datetime.now().strftime('%d-%m-%Y %H:%M')
    ws['B5'].font = normal_font
    ws['B5'].border = border
    ws.merge_cells('B5:D5')

    ws.append([])  # row 6

    # Row 7: column headers
    if is_honda:
        hdrs = ['Part No.', 'Code', 'Description', 'Vehicle', 'Unit', 'Price (₹)', 'Qty']
    else:
        hdrs = ['Part No. (AS)', 'Part No. (SAI)', 'Description', 'Vehicle', 'Colour', 'MRP (₹)', 'Qty']
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = hdr_font
        c.fill = navy
        c.border = border
        c.alignment = center
    ws.row_dimensions[7].height = 22

    # Data rows start at row 8
    for i, item in enumerate(items):
        row = 8 + i
        fill = alt_fill if i % 2 == 0 else white_fill
        vals = [
            item.get('as_part_number', ''),
            item.get('sai_part_number', ''),
            item.get('description', ''),
            item.get('vehicle', ''),
            item.get('colour', ''),
            item.get('mrp', ''),
            item.get('qty', 0),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            c.fill = fill
            c.font = normal_font
            c.alignment = center if col in (1, 2, 6, 7) else left_wrap
        ws.row_dimensions[row].height = 18

    # Column widths
    for col, w in zip('ABCDEFG', [18, 18, 42, 28, 18, 10, 7]):
        ws.column_dimensions[col].width = w

    return wb


def _send_email(config, firm_name, contact_number, items, filepath, fname):
    order_email = config['order_email']
    smtp_user   = config['smtp_user']

    # Always CC the sender account so there's a copy in Gmail Sent/Inbox
    recipients = [order_email]
    if smtp_user and smtp_user != order_email:
        recipients.append(smtp_user)

    msg = MIMEMultipart()
    msg['From']    = smtp_user
    msg['To']      = order_email
    msg['Cc']      = smtp_user
    msg['Subject'] = f"New Order – {firm_name} | {datetime.now().strftime('%d %b %Y %H:%M')}"

    body = (
        f"A new order has been placed via the Rupani Automobiles Order Portal.\n\n"
        f"Firm Name   : {firm_name}\n"
        f"Contact     : {contact_number}\n"
        f"Total Items : {len(items)}\n"
        f"Order Date  : {datetime.now().strftime('%d-%m-%Y %H:%M')}\n\n"
        f"Please find the order details attached."
    )
    msg.attach(MIMEText(body, 'plain'))

    with open(filepath, 'rb') as fh:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(fh.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
        msg.attach(part)

    with smtplib.SMTP(config['smtp_host'], int(config['smtp_port'])) as s:
        s.starttls()
        s.login(smtp_user, config['smtp_pass'])
        s.sendmail(smtp_user, recipients, msg.as_string())


def _wa_post(endpoint, body):
    url = f"{WA_BASE}/{endpoint}/{WA_TOKEN}"
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode(),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    resp = urllib.request.urlopen(req, timeout=20)
    result = json.loads(resp.read())
    print(f"[WA] {endpoint}: {result}", flush=True)
    return result


def _send_whatsapp(firm_name, contact_number, items, filepath, fname, portal='aerostar'):
    """Send order summary + Excel to WhatsApp group via Green API."""
    now       = datetime.now().strftime('%d %b %Y, %I:%M %p')
    total_qty = sum(i.get('qty', 0) for i in items)
    total_mrp = sum((i.get('mrp') or 0) * i.get('qty', 0) for i in items)
    tag = 'Honda Order' if portal == 'honda' else 'Order'

    lines = [
        f"🛒 *New {tag} — {firm_name}*",
        f"📱 {contact_number}",
        f"📦 {len(items)} item(s)  |  Qty: {total_qty}  |  ₹{total_mrp:,.0f}",
        f"🕐 {now}", "",
    ]
    for it in items:
        desc = (it.get('description') or '')[:50]
        lines.append(f"• {it.get('as_part_number','')}  {desc}  ×{it.get('qty',0)}")

    # 1. Send text summary
    _wa_post("sendMessage", {"chatId": WA_GROUP_ID, "message": "\n".join(lines)})

    # 2. Send Excel file as base64 upload
    import base64
    with open(filepath, 'rb') as fh:
        b64 = base64.b64encode(fh.read()).decode()

    _wa_post("sendFileByUpload", {
        "chatId":   WA_GROUP_ID,
        "fileName": fname,
        "caption":  f"📄 {firm_name} — {now}",
        "file":     f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
    })
    print(f"[WA] Order sent for {firm_name}", flush=True)


def _log_to_gsheet(config, firm_name, contact_number, items, email_sent):
    if not _GSPREAD_AVAILABLE:
        print("[GSHEET] gspread not available", flush=True)
        return
    sheet_id = config.get('gsheet_id', '').strip()
    if not sheet_id:
        print("[GSHEET] No sheet ID configured", flush=True)
        return

    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
    ]

    # Load credentials: env var takes priority over file
    creds = None
    creds_b64 = os.environ.get('GOOGLE_CREDENTIALS_B64', '')
    if creds_b64:
        import base64, io
        creds_json = base64.b64decode(creds_b64).decode()
        import google.oauth2.service_account as sa
        creds = sa.Credentials.from_service_account_info(
            json.loads(creds_json), scopes=scopes)
        print("[GSHEET] Using credentials from env var", flush=True)
    else:
        creds_file = os.path.join(BASE_DIR, 'google_credentials.json')
        if not os.path.exists(creds_file):
            print("[GSHEET] No credentials file and no GOOGLE_CREDENTIALS_B64 env var", flush=True)
            return
        creds = SACredentials.from_service_account_file(creds_file, scopes=scopes)
        print("[GSHEET] Using credentials from file", flush=True)

    try:
        gc          = gspread.authorize(creds)
        spreadsheet = gc.open_by_key(sheet_id)

        # ── master Orders sheet ──────────────────────────
        try:
            master = spreadsheet.worksheet('Orders')
        except gspread.WorksheetNotFound:
            master = spreadsheet.add_worksheet(title='Orders', rows=1000, cols=8)
            master.append_row(
                ['Date & Time', 'Firm Name', 'Contact', 'Items', 'Total Qty', 'MRP Total (₹)', 'Email Sent'],
                value_input_option='USER_ENTERED'
            )
            master.format('A1:G1', {
                'textFormat': {'bold': True},
                'backgroundColor': {'red': 0.106, 'green': 0.165, 'blue': 0.290}
            })

        # ── individual order sheet tab ───────────────────
        now       = datetime.now()
        tab_name  = f"{now.strftime('%d%m%y_%H%M')}_{firm_name[:18].strip()}"
        tab_name  = ''.join(c for c in tab_name if c not in r'\/*?[]')[:100]

        # avoid duplicate tab names
        existing  = [ws.title for ws in spreadsheet.worksheets()]
        suffix    = 0
        base_name = tab_name
        while tab_name in existing:
            suffix   += 1
            tab_name  = f"{base_name}_{suffix}"

        order_ws = spreadsheet.add_worksheet(title=tab_name, rows=len(items) + 15, cols=8)
        order_gid = order_ws.id

        # header info rows
        header_rows = [
            ['Rupani Automobiles — Order Sheet'],
            [],
            ['Firm Name',      firm_name],
            ['Contact',        contact_number],
            ['Date & Time',    now.strftime('%d-%m-%Y %H:%M')],
            ['Total Items',    len(items)],
            [],
            ['Part No. (AS)', 'Part No. (SAI)', 'Description', 'Vehicle', 'Colour', 'MRP (₹)', 'Qty'],
        ]
        order_ws.update(range_name='A1', values=header_rows, value_input_option='USER_ENTERED')

        data_rows = [[
            i.get('as_part_number', ''),
            i.get('sai_part_number', ''),
            i.get('description', ''),
            i.get('vehicle', ''),
            i.get('colour', ''),
            i.get('mrp', ''),
            i.get('qty', 0),
        ] for i in items]
        if data_rows:
            order_ws.update(range_name='A9', values=data_rows, value_input_option='USER_ENTERED')

        # ── append row to master with hyperlink ──────────
        total_qty = sum(i.get('qty', 0) for i in items)
        total_mrp = sum((i.get('mrp') or 0) * i.get('qty', 0) for i in items)
        hyperlink  = f'=HYPERLINK("#gid={order_gid}","{firm_name.replace(chr(34), "")}")'

        master.append_row(
            [now.strftime('%d-%m-%Y %H:%M'), hyperlink, contact_number,
             len(items), total_qty, total_mrp, 'Yes' if email_sent else 'No'],
            value_input_option='USER_ENTERED'
        )
    except Exception as e:
        print(f"[GSheet] Error: {e}")


# ── Vehicle lookup helpers ────────────────────────────────────────

# Maps raw API maker strings → catalog brand names
# Surepass returns full legal company names — map all known variants
_BRAND_MAP = {
    # Hero variants (Surepass returns "HERO MOTOCORP LTD")
    'HERO':                          'Hero',
    'HERO MOTOCORP':                 'Hero',
    'HERO MOTOCORP LTD':             'Hero',
    'HERO HONDA':                    'Hero',
    'HERO HONDA MOTORS':             'Hero',
    'HERO HONDA MOTORS LTD':         'Hero',
    # Honda variants — Surepass returns "HONDA MOTORCYCLE & SCOOTER INDIA PVT LTD"
    'HONDA':                                         'Honda',
    'HONDA MOTORCYCLE':                              'Honda',
    'HONDA SCOOTER':                                 'Honda',
    'HONDA MOTORCYCLE AND SCOOTER':                  'Honda',
    'HONDA MOTORCYCLE AND SCOOTER INDIA':            'Honda',
    'HONDA MOTORCYCLE AND SCOOTER INDIA PVT LTD':   'Honda',
    'HONDA MOTORCYCLE AND SCOOTER INDIA PVT. LTD.': 'Honda',
    'HONDA MOTORCYCLE & SCOOTER INDIA PVT LTD':     'Honda',
    'HONDA MOTORCYCLE & SCOOTER INDIA PVT. LTD.':   'Honda',
    'HONDA MOTORCYCLE & SCOOTER':                    'Honda',
    # Bajaj
    'BAJAJ':                         'Bajaj',
    'BAJAJ AUTO':                    'Bajaj',
    'BAJAJ AUTO LTD':                'Bajaj',
    'BAJAJ AUTO LIMITED':            'Bajaj',
    # TVS
    'TVS':                           'Tvs',
    'TVS MOTOR':                     'Tvs',
    'TVS MOTOR COMPANY':             'Tvs',
    'TVS MOTOR COMPANY LTD':         'Tvs',
    'TVS MOTOR CO LTD':              'Tvs',
    # Suzuki
    'SUZUKI':                        'Suzuki',
    'SUZUKI MOTORCYCLE':             'Suzuki',
    'SUZUKI MOTORCYCLE INDIA':       'Suzuki',
    'SUZUKI MOTORCYCLE INDIA PVT LTD': 'Suzuki',
    # Yamaha
    'YAMAHA':                        'Yamaha',
    'INDIA YAMAHA MOTOR':            'Yamaha',
    'INDIA YAMAHA MOTOR PVT LTD':    'Yamaha',
    # Royal Enfield
    'ROYAL ENFIELD':                 'Royal Enfield',
    'ROYAL ENFIELD MOTORS':          'Royal Enfield',
    # Others
    'KTM':                           'Ktm',
    'KTM AG':                        'Ktm',
    'MAHINDRA':                      'Mahindra',
    'MAHINDRA TWO WHEELERS':         'Mahindra',
}


def _normalise_brand(make):
    """Map raw API maker string to catalog brand name."""
    if not make:
        return ''
    # Normalise: uppercase, collapse whitespace, replace & with AND
    upper = re.sub(r'\s+', ' ', make.upper().strip())
    upper_and = upper.replace(' & ', ' AND ')   # "HONDA MOTORCYCLE & SCOOTER" → "HONDA MOTORCYCLE AND SCOOTER"

    # 1. Exact match
    for key_upper in (upper, upper_and):
        if key_upper in _BRAND_MAP:
            return _BRAND_MAP[key_upper]

    # 2. Partial / contains match (longest key first to avoid Hero matching Hero Honda)
    for key, val in sorted(_BRAND_MAP.items(), key=lambda x: -len(x[0])):
        if key in upper or key in upper_and:
            return val

    # 3. Title-case fallback
    return make.title()


def _fetch_vehicle_info(reg_number):
    """
    Fetch vehicle details via Surepass RC V2 API.
    Returns dict {make, model, year, fuel_type, registration_date, owner_name, colour} or None.

    Surepass RC V2 docs: https://app.surepass.app/docs/kyc
    Endpoint: POST https://sandbox.surepass.app/api/v1/rc/rc-full
    Body:     {"id_number": "MH12AB1234"}
    Auth:     Authorization: Bearer <SUREPASS_TOKEN>

    Response data fields used:
      maker_description  → manufacturer (e.g. "HERO MOTOCORP LTD")
      maker_model        → model name   (e.g. "SPLENDOR PLUS")
      manufacturing_date → mfg year
      fuel_type          → PETROL / DIESEL / ELECTRIC
      registration_date  → date of first registration
      owner_name         → registered owner
      color              → vehicle colour
      vehicle_class_desc → Motor Cycle / Scooter etc.
    """
    if not SUREPASS_TOKEN:
        return None   # caller shows "API not configured" message

    try:
        payload = json.dumps({'id_number': reg_number, 'enrich': False}).encode('utf-8')
        req = urllib.request.Request(
            SUREPASS_RC_URL,
            data=payload,
            headers={
                'Authorization':  f'Bearer {SUREPASS_TOKEN}',
                'Content-Type':   'application/json',
                'Accept':         'application/json',
                'User-Agent':     'RupaniOrderPortal/1.0',
            },
            method='POST',
        )
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw = json.loads(resp.read().decode())

        print(f"[Surepass] RC lookup {reg_number}: status={raw.get('status_code')} "
              f"success={raw.get('success')}", flush=True)

        if not raw.get('success'):
            msg = raw.get('message', 'Unknown error')
            print(f"[Surepass] Error: {msg}", flush=True)
            return None

        d = raw.get('data') or {}

        make  = (d.get('maker_description') or '').strip()
        model = (d.get('maker_model')       or '').strip()
        year  = str(d.get('manufacturing_date') or d.get('registration_date') or '')[:4].strip()
        fuel  = (d.get('fuel_type')          or '').strip().title()
        reg_date = (d.get('registration_date') or '').strip()
        owner = (d.get('owner_name')         or '').strip().title()
        colour = (d.get('color')             or '').strip().title()
        veh_class = (d.get('vehicle_class_desc') or '').strip()

        return {
            'make':              make,
            'model':             model,
            'year':              year,
            'fuel_type':         fuel,
            'registration_date': reg_date,
            'owner_name':        owner,
            'colour':            colour,
            'vehicle_class':     veh_class,
            'rc_number':         reg_number,
            # extra fields used by the chassis->model resolver
            'chassis':           (d.get('vehicle_chasi_number') or '').strip(),
            'maker_model':       model,
            'cubic_capacity':    (d.get('cubic_capacity') or '').strip(),
            'norms_type':        (d.get('norms_type') or '').strip(),
        }
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')
        print(f"[Surepass] HTTP {e.code} for {reg_number}: {body[:200]}", flush=True)
        return None
    except Exception as e:
        print(f"[Surepass] Error fetching {reg_number}: {e}", flush=True)
        return None


def _match_parts_to_vehicle(make, model):
    """
    Return list of product IDs from the catalog that match the brand + model.
    Uses fuzzy matching (difflib) for model name comparison.
    """
    brand = _normalise_brand(make)
    if not brand and not model:
        return []

    products = load_products()

    # Prepare model keywords (strip common noise words)
    noise = {'BS6', 'BS4', 'BS-6', 'BS-4', 'FI', 'STD', 'DLX', 'DELUXE', 'DRUM', 'DISC'}
    model_upper = model.upper()
    model_words = [w for w in model_upper.split() if w not in noise and len(w) > 1]

    matched = []
    for p in products:
        # Brand must match (when we have brand info)
        if brand and p.get('brand', '').strip().lower() != brand.lower():
            continue

        vehicle_str = (p.get('vehicle') or '').upper()

        if not model_words:
            matched.append(p['id'])
            continue

        # Check if any model keyword appears in the vehicle string
        keyword_hits = sum(1 for w in model_words if w in vehicle_str)
        if keyword_hits == 0:
            continue

        # Fuzzy ratio check to avoid false positives on short keywords
        ratio = difflib.SequenceMatcher(None, model_upper, vehicle_str).ratio()
        if keyword_hits >= 1 and ratio >= 0.3:
            matched.append(p['id'])
        elif keyword_hits >= 2:
            matched.append(p['id'])

    return matched


@app.route('/api/vehicle-lookup', methods=['POST'])
def vehicle_lookup():
    data       = request.json or {}
    reg_number = (data.get('reg_number') or '').strip().upper()
    reg_number = reg_number.replace(' ', '').replace('-', '')

    if not reg_number or len(reg_number) < 6:
        return jsonify({'success': False, 'error': 'Enter a valid registration number'})

    # Check if API is configured
    if not SUREPASS_TOKEN:
        return jsonify({
            'success': False,
            'error':   (
                'Vehicle lookup is not configured. '
                'Set the SUREPASS_TOKEN environment variable in Railway.'
            )
        })

    vehicle_info = _fetch_vehicle_info(reg_number)

    if not vehicle_info:
        return jsonify({
            'success': False,
            'error':   (
                'Could not find details for this registration number. '
                'Please check the number and try again.'
            )
        })

    make  = vehicle_info.get('make', '').strip()
    model = vehicle_info.get('model', '').strip()

    matched_parts = _match_parts_to_vehicle(make, model)
    brand         = _normalise_brand(make)

    return jsonify({
        'success':     True,
        'vehicle':     vehicle_info,
        'parts_count': len(matched_parts),
        'brand':       brand,
        'model':       model,
    })


# ── Simple self-test (run with: python app.py --test) ────────────
def _run_tests():
    import sys
    errors = []

    # Test 1: _normalise_brand
    result = _normalise_brand("HERO MOTOCORP LTD")
    expected = "Hero"
    if result != expected:
        errors.append(f"FAIL _normalise_brand('HERO MOTOCORP LTD'): got {result!r}, expected {expected!r}")
    else:
        print(f"PASS _normalise_brand('HERO MOTOCORP LTD') -> {result!r}")

    # Test 2: _normalise_brand Honda
    result2 = _normalise_brand("HONDA MOTORCYCLE AND SCOOTER")
    if result2 != "Honda":
        errors.append(f"FAIL _normalise_brand Honda: got {result2!r}")
    else:
        print(f"PASS _normalise_brand('HONDA MOTORCYCLE AND SCOOTER') -> {result2!r}")

    # Test 3: _match_parts_to_vehicle – only meaningful if catalog is loaded
    try:
        matched = _match_parts_to_vehicle("HONDA MOTORCYCLE AND SCOOTER", "ACTIVA 6G")
        if len(matched) > 0:
            print(f"PASS _match_parts_to_vehicle('HONDA...', 'ACTIVA 6G') -> {len(matched)} parts")
        else:
            print(f"INFO _match_parts_to_vehicle('HONDA...', 'ACTIVA 6G') -> 0 parts (catalog may not contain Activa)")
    except Exception as e:
        errors.append(f"FAIL _match_parts_to_vehicle raised: {e}")

    if errors:
        print("\nFAILURES:")
        for e in errors:
            print(" ", e)
        sys.exit(1)
    else:
        print("\nAll tests passed.")


# establish the Honda parts baseline on startup (safe/idempotent across workers)
try:
    sync_honda_registry()
except Exception as _e:
    print(f"[HONDA REGISTRY STARTUP] {_e}", flush=True)


if __name__ == '__main__':
    import sys
    if '--test' in sys.argv:
        _run_tests()
    else:
        app.run(debug=True, port=5001, host='0.0.0.0')
