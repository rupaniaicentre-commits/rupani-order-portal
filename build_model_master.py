#!/usr/bin/env python3
"""Build model_master_review.xlsx for two-wheeler parts portal."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Master data ────────────────────────────────────────────────────────────
# Columns: Brand, Official_Model_Name, Category, Status, BS_Norm, Notes

MODELS = []

def add(brand, name, category, status, bs_norm, notes=""):
    MODELS.append({
        "Brand": brand,
        "Official_Model_Name": name,
        "Category": category,
        "Status": status,
        "BS_Norm": bs_norm,
        "Notes": notes,
    })

# ════════════════════════════════════════════════════════════
# HERO MOTOCORP
# ════════════════════════════════════════════════════════════
H = "Hero"

# Current Motorcycles
add(H, "Splendor Plus", "Motorcycle", "Active", "OBD2")
add(H, "Splendor Plus Xtec", "Motorcycle", "Active", "OBD2", "Connected Bluetooth edition")
add(H, "Splendor Plus OBD2", "Motorcycle", "Active", "OBD2", "OBD2 compliant variant")
add(H, "Super Splendor", "Motorcycle", "Active", "OBD2")
add(H, "Super Splendor Xtec", "Motorcycle", "Active", "OBD2", "Bluetooth connectivity")
add(H, "Glamour", "Motorcycle", "Active", "OBD2")
add(H, "Glamour Xtec", "Motorcycle", "Active", "OBD2", "Bluetooth connectivity variant")
add(H, "Glamour Digital", "Motorcycle", "Active", "BS6", "Full digital console variant")
add(H, "HF Deluxe", "Motorcycle", "Active", "OBD2")
add(H, "HF Deluxe Eco", "Motorcycle", "Active", "BS6", "i3S idle stop-start system")
add(H, "Passion Pro", "Motorcycle", "Active", "OBD2")
add(H, "Passion Pro i3S", "Motorcycle", "Active", "BS6", "i3S idle stop-start variant")
add(H, "Passion Xpro", "Motorcycle", "Active", "BS6")
add(H, "Xtreme 160R", "Motorcycle", "Active", "OBD2")
add(H, "Xtreme 160R 4V", "Motorcycle", "Active", "OBD2", "4-valve engine variant")
add(H, "Xtreme 200R", "Motorcycle", "Active", "BS6")
add(H, "Xtreme 200S", "Motorcycle", "Active", "BS6", "Semi-faired sportbike")
add(H, "Xpulse 200", "Motorcycle", "Active", "BS6", "Adventure tourer")
add(H, "Xpulse 200 4V", "Motorcycle", "Active", "OBD2", "4-valve adventure variant")
add(H, "Xpulse 200T", "Motorcycle", "Active", "BS6", "Touring variant of Xpulse")
add(H, "Mavrick 440", "Motorcycle", "Active", "OBD2", "Premium Neo-Retro roadster; also Mavrick 440")

# Current Scooters
add(H, "Maestro Edge 110", "Scooter", "Active", "OBD2")
add(H, "Maestro Edge 125", "Scooter", "Active", "OBD2")
add(H, "Destini 125", "Scooter", "Active", "OBD2")
add(H, "Destini Prime", "Scooter", "Active", "OBD2", "Premium variant of Destini 125")
add(H, "Pleasure Plus", "Scooter", "Active", "OBD2")
add(H, "Pleasure Plus Xtec", "Scooter", "Active", "OBD2", "Bluetooth connectivity")
add(H, "Xoom 110", "Scooter", "Active", "OBD2")
add(H, "Xoom 125", "Scooter", "Active", "OBD2")

# Current Electric
add(H, "Vida V1", "Electric", "Active", "Electric", "Hero's electric scooter; Vida V1 Pro / Plus variants")
add(H, "Vida V2", "Electric", "Active", "Electric", "2nd gen Vida electric scooter")

# Discontinued Motorcycles
add(H, "CD 100", "Motorcycle", "Discontinued", "Pre-BS4", "Original iconic commuter")
add(H, "CD 100 SS", "Motorcycle", "Discontinued", "Pre-BS4", "Self-start variant of CD100")
add(H, "CD Dawn", "Motorcycle", "Discontinued", "BS4")
add(H, "CD Deluxe", "Motorcycle", "Discontinued", "BS4")
add(H, "CBZ", "Motorcycle", "Discontinued", "Pre-BS4", "Original CBZ sports bike")
add(H, "CBZ Xtreme", "Motorcycle", "Discontinued", "BS4")
add(H, "Hunk", "Motorcycle", "Discontinued", "BS4")
add(H, "Karizma", "Motorcycle", "Discontinued", "BS4", "Original 223cc Karizma")
add(H, "Karizma R", "Motorcycle", "Discontinued", "BS4")
add(H, "Karizma ZMR", "Motorcycle", "Discontinued", "BS4", "Fully-faired variant")
add(H, "Xtreme Sports", "Motorcycle", "Discontinued", "BS4")
add(H, "Achiever", "Motorcycle", "Discontinued", "BS4")
add(H, "Ignitor", "Motorcycle", "Discontinued", "BS4")
add(H, "Impulse", "Motorcycle", "Discontinued", "BS4", "Off-road dual sport")
add(H, "Ambition", "Motorcycle", "Discontinued", "Pre-BS4")
add(H, "Passion", "Motorcycle", "Discontinued", "Pre-BS4", "Original Passion")
add(H, "Passion Plus", "Motorcycle", "Discontinued", "Pre-BS4")
add(H, "Splendor", "Motorcycle", "Discontinued", "Pre-BS4", "Original Hero Honda Splendor")
add(H, "Splendor NXG", "Motorcycle", "Discontinued", "BS4")
add(H, "Splendor i-Smart", "Motorcycle", "Discontinued", "BS4", "i3S auto-start stop")
add(H, "Splendor Pro", "Motorcycle", "Discontinued", "BS4")
add(H, "HF Dawn", "Motorcycle", "Discontinued", "BS4")

# Discontinued Scooters
add(H, "Maestro Edge", "Scooter", "Discontinued", "BS4", "Original 110cc Maestro Edge")


# ════════════════════════════════════════════════════════════
# HONDA MOTORCYCLE & SCOOTER INDIA
# ════════════════════════════════════════════════════════════
HO = "Honda"

# Current Motorcycles
add(HO, "CB Shine", "Motorcycle", "Active", "OBD2")
add(HO, "CB Shine SP", "Motorcycle", "Active", "BS6", "Self-start plus variant")
add(HO, "SP 125", "Motorcycle", "Active", "OBD2")
add(HO, "SP 160", "Motorcycle", "Active", "OBD2")
add(HO, "Livo", "Motorcycle", "Active", "OBD2")
add(HO, "CD 110 Dream", "Motorcycle", "Active", "OBD2")
add(HO, "CD 110 Dream DX", "Motorcycle", "Active", "OBD2", "Deluxe variant")
add(HO, "CB Unicorn", "Motorcycle", "Active", "BS6")
add(HO, "CB Unicorn 160", "Motorcycle", "Discontinued", "BS4", "Replaced by CB200X / Hornet")
add(HO, "CB Hornet 160R", "Motorcycle", "Discontinued", "BS6", "Discontinued post-2021")
add(HO, "CB200X", "Motorcycle", "Active", "OBD2", "Adventure crossover")
add(HO, "CB300R", "Motorcycle", "Active", "BS6", "Neo-sports cafe")
add(HO, "CB350", "Motorcycle", "Active", "OBD2", "H'Ness CB350 premium retro")
add(HO, "CB350RS", "Motorcycle", "Active", "OBD2", "Scrambler variant of CB350")
add(HO, "CB500X", "Motorcycle", "Active", "BS6", "Mid-size adventure tourer (CBU)")
add(HO, "CBR150R", "Motorcycle", "Discontinued", "BS4")
add(HO, "CBR250R", "Motorcycle", "Discontinued", "Pre-BS4")
add(HO, "CBR650R", "Motorcycle", "Active", "BS6", "Fully-faired supersport (CBU)")
add(HO, "XBlade", "Motorcycle", "Active", "BS6")
add(HO, "NX500", "Motorcycle", "Active", "OBD2", "Launched 2023; mid-size adventure")
add(HO, "CB650R", "Motorcycle", "Active", "BS6", "Neo-sports cafe (CBU)")
add(HO, "Africa Twin", "Motorcycle", "Active", "BS6", "CRF1100L adventure tourer (CBU)")

# Current Scooters
add(HO, "Activa 6G", "Scooter", "Active", "OBD2", "Current-gen Activa 110")
add(HO, "Activa 125", "Scooter", "Active", "OBD2", "BS6 OBD2 variant launched 2023")
add(HO, "Activa 125 Premium Edition", "Scooter", "Active", "BS6", "Special edition trim")
add(HO, "Activa e:", "Scooter", "Active", "Electric", "Electric variant of Activa; launched 2024")
add(HO, "Dio", "Motorcycle", "Discontinued", "BS4", "110cc scooter; replaced by Dio 125")
add(HO, "Dio 125", "Scooter", "Active", "OBD2")
add(HO, "Grazia 125", "Scooter", "Active", "BS6")
add(HO, "Vision 110", "Scooter", "Active", "OBD2")
add(HO, "Cliq", "Scooter", "Discontinued", "BS4", "Budget scooter with CBS")
add(HO, "Navi", "Scooter", "Discontinued", "BS4", "Crossover scooter")
add(HO, "Aviator", "Scooter", "Discontinued", "Pre-BS4")

# Discontinued Motorcycles
add(HO, "Dream Yuga", "Motorcycle", "Discontinued", "BS4")
add(HO, "Dream Neo", "Motorcycle", "Discontinued", "BS4")
add(HO, "Stunner CBF", "Motorcycle", "Discontinued", "Pre-BS4")
add(HO, "Twister", "Motorcycle", "Discontinued", "Pre-BS4")

# Discontinued Scooters
add(HO, "Activa", "Scooter", "Discontinued", "Pre-BS4", "Original Activa variants pre-3G")
add(HO, "Activa i", "Scooter", "Discontinued", "BS4", "Smaller-wheel economy variant")
add(HO, "Activa 3G", "Scooter", "Discontinued", "BS4")
add(HO, "Activa 4G", "Scooter", "Discontinued", "BS4")
add(HO, "Activa 5G", "Scooter", "Discontinued", "BS4")
add(HO, "Activa 125 BS6", "Scooter", "Discontinued", "BS6", "Pre-OBD2 BS6 variant; replaced by OBD2")


# ════════════════════════════════════════════════════════════
# TVS MOTOR
# ════════════════════════════════════════════════════════════
TV = "TVS"

# Current Motorcycles
add(TV, "Apache RTR 160", "Motorcycle", "Active", "OBD2")
add(TV, "Apache RTR 160 2V", "Motorcycle", "Active", "OBD2", "2-valve engine version")
add(TV, "Apache RTR 160 4V", "Motorcycle", "Active", "OBD2", "4-valve flagship 160cc")
add(TV, "Apache RTR 180", "Motorcycle", "Active", "OBD2")
add(TV, "Apache RTR 200 4V", "Motorcycle", "Active", "OBD2")
add(TV, "Apache RR 310", "Motorcycle", "Active", "OBD2", "Fully-faired 312cc supersport")
add(TV, "Raider 125", "Motorcycle", "Active", "OBD2")
add(TV, "Star City Plus", "Motorcycle", "Active", "OBD2")
add(TV, "Sport", "Motorcycle", "Active", "OBD2", "Entry-level commuter")
add(TV, "Radeon", "Motorcycle", "Active", "OBD2")
add(TV, "Ronin", "Motorcycle", "Active", "OBD2", "225cc neo-retro roadster launched 2022")

# Discontinued Motorcycles
add(TV, "Metro 100", "Motorcycle", "Discontinued", "Pre-BS4")
add(TV, "Victor", "Motorcycle", "Discontinued", "BS4", "Also sold as Victor GLX / Edge")

# Current Scooters
add(TV, "Jupiter", "Scooter", "Active", "OBD2")
add(TV, "Jupiter 125", "Scooter", "Active", "OBD2")
add(TV, "Ntorq 125", "Scooter", "Active", "OBD2", "Sports scooter with connected features")
add(TV, "Scooty Pep Plus", "Scooter", "Active", "OBD2")
add(TV, "Scooty Zest 110", "Scooter", "Active", "OBD2")
add(TV, "iQube Electric", "Electric", "Active", "Electric", "Electric scooter; ST / S variants")
add(TV, "iQube S", "Electric", "Active", "Electric", "iQube S variant with larger battery")

# Mopeds
add(TV, "XL100", "Moped", "Active", "OBD2")
add(TV, "XL100 Comfort", "Moped", "Active", "OBD2")
add(TV, "XL100 Heavy Duty", "Moped", "Active", "OBD2")
add(TV, "XL Super", "Moped", "Active", "OBD2", "Upgraded XL moped")


# ════════════════════════════════════════════════════════════
# BAJAJ AUTO
# ════════════════════════════════════════════════════════════
BA = "Bajaj"

# Current Motorcycles
add(BA, "CT 100", "Motorcycle", "Active", "OBD2")
add(BA, "CT 110", "Motorcycle", "Active", "OBD2")
add(BA, "CT 110X", "Motorcycle", "Active", "OBD2", "Off-road inspired variant")
add(BA, "Platina 100", "Motorcycle", "Active", "OBD2")
add(BA, "Platina 110", "Motorcycle", "Active", "OBD2")
add(BA, "Pulsar 125", "Motorcycle", "Active", "OBD2")
add(BA, "Pulsar 150", "Motorcycle", "Active", "OBD2")
add(BA, "Pulsar 160NS", "Motorcycle", "Active", "OBD2", "Naked street 160cc")
add(BA, "Pulsar 180", "Motorcycle", "Active", "OBD2")
add(BA, "Pulsar 180F", "Motorcycle", "Active", "OBD2", "Semi-faired 180cc")
add(BA, "Pulsar 200NS", "Motorcycle", "Active", "OBD2")
add(BA, "Pulsar 220F", "Motorcycle", "Active", "OBD2", "Iconic fully-faired 220cc")
add(BA, "Pulsar 250", "Motorcycle", "Active", "OBD2", "250cc naked street")
add(BA, "Pulsar N160", "Motorcycle", "Active", "OBD2", "N-series 160cc")
add(BA, "Pulsar N250", "Motorcycle", "Active", "OBD2", "N-series 250cc naked")
add(BA, "Pulsar F250", "Motorcycle", "Active", "OBD2", "F-series 250cc semi-faired")
add(BA, "Pulsar RS200", "Motorcycle", "Active", "OBD2", "Fully-faired 200cc")
add(BA, "Avenger Street 160", "Motorcycle", "Active", "OBD2", "Cruiser")
add(BA, "Avenger Cruise 220", "Motorcycle", "Active", "OBD2", "Cruiser")
add(BA, "Dominar 250", "Motorcycle", "Active", "OBD2", "Performance tourer")
add(BA, "Dominar 400", "Motorcycle", "Active", "OBD2", "Flagship tourer")

# Current Commuter/Entry
add(BA, "Discover 100", "Motorcycle", "Active", "OBD2")
add(BA, "Discover 110", "Motorcycle", "Active", "OBD2")
add(BA, "Discover 125", "Motorcycle", "Active", "OBD2")

# Discontinued
add(BA, "Caliber", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "Caliber 115", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "Caliber Croma", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "Wind 125", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "Spirit", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "Classic 350 (Bajaj)", "Motorcycle", "Discontinued", "Pre-BS4", "Old Bajaj Classic; not to be confused with RE Classic")
add(BA, "KB100", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "KB125", "Motorcycle", "Discontinued", "Pre-BS4")
add(BA, "XCD 125", "Motorcycle", "Discontinued", "BS4")
add(BA, "XCD 135", "Motorcycle", "Discontinued", "BS4")
add(BA, "Pulsar 135 LS", "Motorcycle", "Discontinued", "BS4", "Lightest Pulsar; LS = Light Sport")
add(BA, "Pulsar 150 DTSi", "Motorcycle", "Discontinued", "BS4", "Early DTS-i twin spark variant")
add(BA, "Pulsar 180 DTSi", "Motorcycle", "Discontinued", "BS4")
add(BA, "Pulsar 200 NS (BS4)", "Motorcycle", "Discontinued", "BS4", "Earlier NS200 before current OBD2")
add(BA, "Discover 150", "Motorcycle", "Discontinued", "BS4")
add(BA, "Boxer 100", "Motorcycle", "Discontinued", "BS4", "Rural/semi-urban commuter")
add(BA, "Boxer 150", "Motorcycle", "Discontinued", "BS4")
add(BA, "Avenger 220 DTSi", "Motorcycle", "Discontinued", "BS4", "Earlier Avenger 220 variant")


# ════════════════════════════════════════════════════════════
# SUZUKI MOTORCYCLE INDIA
# ════════════════════════════════════════════════════════════
SU = "Suzuki"

# Current Motorcycles
add(SU, "Gixxer", "Motorcycle", "Active", "OBD2", "150cc naked street")
add(SU, "Gixxer 250", "Motorcycle", "Active", "OBD2", "250cc naked")
add(SU, "Gixxer SF", "Motorcycle", "Active", "OBD2", "150cc semi-faired")
add(SU, "Gixxer SF 250", "Motorcycle", "Active", "OBD2", "250cc fully-faired")
add(SU, "Intruder 150", "Motorcycle", "Active", "OBD2", "Cruiser 150cc")
add(SU, "V-Strom SX 250", "Motorcycle", "Active", "OBD2", "Adventure tourer 250cc")

# Current Scooters
add(SU, "Access 125", "Scooter", "Active", "OBD2")
add(SU, "Burgman Street 125", "Scooter", "Active", "OBD2", "Maxi-style scooter")
add(SU, "Avenis 125", "Scooter", "Active", "OBD2", "Sporty scooter")

# Discontinued Motorcycles
add(SU, "Hayate", "Motorcycle", "Discontinued", "BS4", "110cc commuter")
add(SU, "Heat", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Fiero", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Samurai", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Shogun", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Max 100", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Max 100R", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Slingshot", "Motorcycle", "Discontinued", "Pre-BS4")
add(SU, "Slingshot Plus", "Motorcycle", "Discontinued", "BS4")
add(SU, "GS150R", "Motorcycle", "Discontinued", "Pre-BS4", "150cc sport naked")

# Discontinued Scooters
add(SU, "Swish 125", "Scooter", "Discontinued", "BS4")
add(SU, "Lets", "Scooter", "Discontinued", "BS4", "Budget lightweight scooter")


# ════════════════════════════════════════════════════════════
# YAMAHA INDIA
# ════════════════════════════════════════════════════════════
YA = "Yamaha"

# Current Motorcycles
add(YA, "FZ-S FI V3", "Motorcycle", "Active", "OBD2", "FZ-S FI 150cc V3.0")
add(YA, "FZS-FI V3", "Motorcycle", "Active", "OBD2", "Also listed as FZS V3; dark knight edition")
add(YA, "FZ-X", "Motorcycle", "Active", "OBD2", "Crossover FZ; launched 2021")
add(YA, "FZ25", "Motorcycle", "Active", "OBD2", "250cc naked street")
add(YA, "FZS25", "Motorcycle", "Active", "OBD2", "FZ25 with more features/color options")
add(YA, "MT-15 V2", "Motorcycle", "Active", "OBD2", "V2 update launched 2022")
add(YA, "R15 V4", "Motorcycle", "Active", "OBD2", "Fully-faired supersport")
add(YA, "R15M", "Motorcycle", "Active", "OBD2", "Monster Energy edition; top-spec R15")
add(YA, "R15S", "Motorcycle", "Active", "OBD2", "Entry R15 variant")
add(YA, "R7", "Motorcycle", "Active", "BS6", "689cc middleweight supersport (CBU)")
add(YA, "MT-03", "Motorcycle", "Active", "BS6", "321cc naked twin (CBU)")
add(YA, "YZF-R3", "Motorcycle", "Active", "BS6", "321cc fully-faired (CBU)")

# Current Scooters
add(YA, "Fascino 125 FI", "Scooter", "Active", "OBD2")
add(YA, "Ray ZR 125 FI", "Scooter", "Active", "OBD2")
add(YA, "Ray ZR Street Rally 125 FI", "Scooter", "Active", "OBD2", "Off-road inspired variant")
add(YA, "Aerox 155", "Scooter", "Active", "OBD2", "Maxi-scooter 155cc")

# Discontinued Motorcycles
add(YA, "RX 100", "Motorcycle", "Discontinued", "Pre-BS4", "Iconic 2-stroke 98cc")
add(YA, "RX 135", "Motorcycle", "Discontinued", "Pre-BS4", "135cc 2-stroke")
add(YA, "RXG", "Motorcycle", "Discontinued", "Pre-BS4", "98cc 2-stroke variant")
add(YA, "Alba", "Motorcycle", "Discontinued", "Pre-BS4")
add(YA, "Enticer", "Motorcycle", "Discontinued", "Pre-BS4", "125cc cruiser")
add(YA, "Gladiator", "Motorcycle", "Discontinued", "Pre-BS4")
add(YA, "Crux", "Motorcycle", "Discontinued", "BS4", "Entry-level commuter")
add(YA, "Libero", "Motorcycle", "Discontinued", "Pre-BS4")
add(YA, "Saluto 125", "Motorcycle", "Discontinued", "BS4", "125cc commuter")
add(YA, "Saluto RX", "Motorcycle", "Discontinued", "BS4", "Spoke-wheel version of Saluto")
add(YA, "SZ-RR", "Motorcycle", "Discontinued", "BS4", "150cc semi-naked")
add(YA, "SZR", "Motorcycle", "Discontinued", "BS4", "150cc variant")
add(YA, "SZ-S", "Motorcycle", "Discontinued", "BS4")
add(YA, "YBR 110", "Motorcycle", "Discontinued", "BS4")
add(YA, "YBR 125", "Motorcycle", "Discontinued", "BS4")
add(YA, "Fazer", "Motorcycle", "Discontinued", "BS4", "150cc half-faired; also Fazer FI V2")
add(YA, "FZ16", "Motorcycle", "Discontinued", "Pre-BS4", "Original FZ 153cc")
add(YA, "FZS (old)", "Motorcycle", "Discontinued", "Pre-BS4", "FZS first generation")
add(YA, "MT-15 V1", "Motorcycle", "Discontinued", "BS6", "V1 replaced by V2 in 2022")
add(YA, "R15 V3", "Motorcycle", "Discontinued", "BS6", "V3 replaced by V4 in 2021")

# Discontinued Scooters
add(YA, "Cygnus Ray", "Scooter", "Discontinued", "Pre-BS4", "Original Cygnus Ray")
add(YA, "Cygnus Ray Z", "Scooter", "Discontinued", "BS4", "Refreshed Cygnus with disc brake")
add(YA, "Ray ZR (BS4)", "Scooter", "Discontinued", "BS4", "Pre-BS6 Ray ZR")

# ─── Sort by Brand then Model Name ──────────────────────────────────────────
MODELS.sort(key=lambda x: (x["Brand"], x["Official_Model_Name"]))

# ─── Build workbook ──────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Model Master ────────────────────────────────────────────────────
ws = wb.active
ws.title = "Model Master"

HEADERS = ["Brand", "Official_Model_Name", "Category", "Status", "BS_Norm", "Notes"]
COL_WIDTHS = [12, 40, 14, 14, 12, 55]

# Style helpers
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="DCE6F1")
PLAIN_FILL = PatternFill("solid", start_color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Status colour map
STATUS_FONTS = {
    "Active": Font(name="Arial", size=10, color="006100"),        # dark green
    "Discontinued": Font(name="Arial", size=10, color="9C0006"),  # dark red
}
STATUS_FILLS = {
    "Active": PatternFill("solid", start_color="C6EFCE"),
    "Discontinued": PatternFill("solid", start_color="FFC7CE"),
}

# Write headers
for col_idx, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=col_idx, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 20

# Write data rows
brand_colors = {}
brand_list = ["Hero", "Honda", "TVS", "Bajaj", "Suzuki", "Yamaha"]
BRAND_FILLS = {
    "Hero":   "E2EFDA",
    "Honda":  "FFF2CC",
    "TVS":    "FCE4D6",
    "Bajaj":  "EBF3FB",
    "Suzuki": "F4CCFF",
    "Yamaha": "FDEBD0",
}

for row_idx, model in enumerate(MODELS, start=2):
    brand = model["Brand"]
    brand_fill = PatternFill("solid", start_color=BRAND_FILLS.get(brand, "FFFFFF"))
    status = model["Status"]

    for col_idx, key in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=model[key])
        cell.border = BORDER

        if col_idx == 4:  # Status column
            cell.font = STATUS_FONTS.get(status, DATA_FONT)
            cell.fill = STATUS_FILLS.get(status, PLAIN_FILL)
            cell.alignment = CENTER
        elif col_idx == 6:  # Notes
            cell.font = Font(name="Arial", size=9, italic=True, color="595959")
            cell.fill = brand_fill
            cell.alignment = LEFT
        elif col_idx in (3, 5):  # Category, BS_Norm
            cell.font = DATA_FONT
            cell.fill = brand_fill
            cell.alignment = CENTER
        else:
            cell.font = DATA_FONT
            cell.fill = brand_fill
            cell.alignment = LEFT

    ws.row_dimensions[row_idx].height = 16

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# ── Sheet 2: Per Brand Count ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Per Brand Count")

BRAND_HEADER_FILL = PatternFill("solid", start_color="1F4E79")

cnt_headers = ["Brand", "Total Models", "Active", "Discontinued",
               "Motorcycle", "Scooter", "Moped", "Electric"]
cnt_widths   = [14, 14, 10, 14, 12, 10, 10, 10]

for col_idx, (hdr, width) in enumerate(zip(cnt_headers, cnt_widths), start=1):
    cell = ws2.cell(row=1, column=col_idx, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = BRAND_HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.row_dimensions[1].height = 20

for row_idx, brand in enumerate(brand_list, start=2):
    brand_models = [m for m in MODELS if m["Brand"] == brand]
    total = len(brand_models)
    active = sum(1 for m in brand_models if m["Status"] == "Active")
    disc   = sum(1 for m in brand_models if m["Status"] == "Discontinued")
    moto   = sum(1 for m in brand_models if m["Category"] == "Motorcycle")
    scoot  = sum(1 for m in brand_models if m["Category"] == "Scooter")
    moped  = sum(1 for m in brand_models if m["Category"] == "Moped")
    elec   = sum(1 for m in brand_models if m["Category"] == "Electric")

    row_data = [brand, total, active, disc, moto, scoot, moped, elec]
    bf = PatternFill("solid", start_color=BRAND_FILLS.get(brand, "FFFFFF"))

    for col_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=10, bold=(col_idx == 1))
        cell.fill = bf
        cell.border = BORDER
        cell.alignment = CENTER if col_idx > 1 else LEFT

    ws2.row_dimensions[row_idx].height = 18

# Totals row
total_row = len(brand_list) + 2
ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
ws2.cell(row=total_row, column=1).border = BORDER
ws2.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="D9D9D9")
ws2.cell(row=total_row, column=1).alignment = CENTER

for col_idx in range(2, len(cnt_headers) + 1):
    col_letter = get_column_letter(col_idx)
    start_row = 2
    end_row = len(brand_list) + 1
    cell = ws2.cell(row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
    cell.font = Font(name="Arial", bold=True)
    cell.fill = PatternFill("solid", start_color="D9D9D9")
    cell.border = BORDER
    cell.alignment = CENTER

ws2.row_dimensions[total_row].height = 18
ws2.freeze_panes = "A2"

# ─── Save ────────────────────────────────────────────────────────────────────
OUT = "/Users/rupaniai/my-new-personal-project/rupani-order-portal/model_master_review.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Total models: {len(MODELS)}")

from collections import Counter
brand_counts = Counter(m["Brand"] for m in MODELS)
for b in brand_list:
    print(f"  {b}: {brand_counts[b]}")
